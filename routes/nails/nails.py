"""
nails.py — Blueprint principal del módulo Nails (salón de uñas)
Cubre: onboarding, dashboard, agenda, clientes, servicios, ventas,
reportes, configuración, galería, catálogo público y uploads a R2.
"""

import re
import os
import io
import json
import uuid
import base64
import threading
import secrets
import requests
import boto3
from botocore.config import Config
from datetime import date, datetime, timedelta
from werkzeug.utils import secure_filename
from flask import (
    Blueprint, render_template, redirect, url_for,
    session, request, flash, current_app, jsonify,
    send_file,
)
from db import get_db_connection
from utils.nails_security import (
    nails_csrf_context,
    nails_csrf_error_response,
    public_booking_rate_limited,
    validate_nails_csrf,
)


nails_bp = Blueprint("nails", __name__, url_prefix="/nails")


# =========================================================
# CONFIGURACIÓN DE CLOUDFLARE R2 PARA NAILS
# Lee las credenciales desde variables de entorno para no
# hardcodear secretos en el código.
# =========================================================

ACCESS_KEY   = os.getenv("R2_ACCESS_KEY")
SECRET_KEY   = os.getenv("R2_SECRET_KEY")
ENDPOINT_URL = os.getenv("R2_ENDPOINT_URL")
BUCKET_NAME  = os.getenv("R2_BUCKET_NAME")
PUBLIC_URL   = os.getenv("R2_PUBLIC_URL")

ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "svg", "webp"}

# Conjuntos de valores válidos para campos de estado/tipo.
# Validar contra estos sets evita SQL con valores arbitrarios.
APPOINTMENT_STATUSES = {"pendiente", "confirmada", "atendida", "cancelada", "no_asistio"}
PAYMENT_METHODS      = {"efectivo", "transferencia", "tarjeta", "mixto", "otro"}
SALE_STATUSES        = {"pendiente", "anticipo", "pagada", "cancelada"}
EXPENSE_CATEGORIES   = {
    "materiales", "renta", "servicios", "sueldos", "comisiones",
    "publicidad", "mantenimiento", "capacitacion", "otros",
}
EXPENSE_FREQUENCIES  = {"semanal", "quincenal", "mensual", "bimestral", "anual"}
EXPENSE_STATUSES     = {"activo", "cancelado"}

# Colores por defecto si el usuario no elige ninguno en el onboarding/config.
DEFAULT_PRIMARY_COLOR   = "#ff6b81"
DEFAULT_SECONDARY_COLOR = "#fff0f3"
DEFAULT_ACCENT_COLOR    = "#2ec4b6"

NAILS_SERVICE_ICON_OPTIONS = [
    ("hand-sparkles", "Manos"),
    ("wand-magic-sparkles", "Diseño"),
    ("bottle-droplet", "Gelish"),
    ("paintbrush", "Acrílico"),
    ("gem", "Pedrería"),
    ("spa", "Spa"),
    ("shoe-prints", "Pedicure"),
    ("droplet", "Efecto"),
    ("heart", "Favorito"),
    ("star", "Especial"),
]
NAILS_SERVICE_ICONS = {icon for icon, _label in NAILS_SERVICE_ICON_OPTIONS}


# =========================================================
# SINGLETON DE CLIENTE S3 (BOTO3 / R2)
# Se usa un lock para que en entornos multihilo (Gunicorn)
# nunca se creen dos instancias simultáneamente.
# BUG CORREGIDO: antes no había lock → race condition posible.
# =========================================================

_s3_client      = None
_s3_client_lock = threading.Lock()


def get_s3_client():
    """
    Devuelve el cliente boto3 configurado para Cloudflare R2.
    Lanza RuntimeError si faltan variables de entorno.
    El singleton se crea una sola vez gracias al threading.Lock.
    """
    global _s3_client

    missing_config = [
        name for name, value in {
            "R2_ACCESS_KEY":   ACCESS_KEY,
            "R2_SECRET_KEY":   SECRET_KEY,
            "R2_ENDPOINT_URL": ENDPOINT_URL,
            "R2_BUCKET_NAME":  BUCKET_NAME,
            "R2_PUBLIC_URL":   PUBLIC_URL,
        }.items()
        if not value
    ]

    if missing_config:
        raise RuntimeError(
            "Faltan variables de entorno para subir imágenes: "
            + ", ".join(missing_config)
        )

    # Doble comprobación dentro del lock para crear el cliente
    # solo una vez, incluso con peticiones concurrentes.
    if _s3_client is None:
        with _s3_client_lock:
            if _s3_client is None:
                _s3_client = boto3.client(
                    service_name="s3",
                    endpoint_url=ENDPOINT_URL,
                    aws_access_key_id=ACCESS_KEY,
                    aws_secret_access_key=SECRET_KEY,
                    region_name="auto",
                    config=Config(signature_version="s3v4"),
                )

    return _s3_client


# =========================================================
# HELPERS DE SANITIZACIÓN Y VALIDACIÓN
# Funciones puras que normalizan y validan datos de formulario
# antes de que toquen la base de datos.
# =========================================================

def allowed_image_file(filename):
    """Devuelve True si la extensión del archivo está en la whitelist."""
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS
    )


def clean_text(value, max_length=None):
    """Elimina espacios al inicio/fin y recorta al largo máximo."""
    text = (value or "").strip()
    if max_length:
        return text[:max_length]
    return text


def is_general_customer_name(value):
    return clean_text(value).lower() == "cliente general"


def default_service_icon_for_category(category_name):
    category = clean_text(category_name).lower()
    if "gel" in category:
        return "bottle-droplet"
    if "acril" in category or "acryl" in category:
        return "paintbrush"
    if "pedi" in category or "pie" in category:
        return "shoe-prints"
    if "extra" in category or "pedrer" in category:
        return "gem"
    if "spa" in category:
        return "spa"
    if "dise" in category or "arte" in category:
        return "wand-magic-sparkles"
    return "hand-sparkles"


def clean_service_icon(value, category_name=None):
    icon = clean_text(value, 60).replace("fa-", "")
    if icon in NAILS_SERVICE_ICONS:
        return icon
    return default_service_icon_for_category(category_name)


def ensure_nails_service_icon_column(cur):
    cur.execute(
        """
        ALTER TABLE nails_services
        ADD COLUMN IF NOT EXISTS service_icon TEXT DEFAULT 'hand-sparkles'
        """
    )


def parse_positive_float(value, default=0.0, max_value=None):
    """
    Convierte montos del formulario a float.
    Nunca devuelve valores negativos; aplica techo opcional.
    """
    try:
        number = float(value or default)
    except (TypeError, ValueError):
        return default

    if number < 0:
        number = 0.0

    if max_value is not None and number > max_value:
        number = max_value

    return number


def parse_positive_int(value, default=0, max_value=None):
    """
    Convierte duraciones/órdenes del formulario a int.
    Nunca devuelve valores negativos; aplica techo opcional.
    """
    try:
        number = int(value or default)
    except (TypeError, ValueError):
        return default

    if number < 0:
        number = 0

    if max_value is not None and number > max_value:
        number = max_value

    return number


def parse_date_value(value):
    """
    Convierte string 'YYYY-MM-DD' a objeto date.
    Devuelve None si el valor está vacío o tiene formato incorrecto.
    """
    if not value:
        return None

    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_time_to_minutes(value):
    """Convierte horas tipo 10:00, 10, 7:30 pm a minutos del día."""
    text = clean_text(value).lower().replace(".", "")
    match = re.fullmatch(r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)?", text)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2) or 0)
    meridian = match.group(3)

    if minute > 59:
        return None

    if meridian:
        if hour < 1 or hour > 12:
            return None
        if meridian == "pm" and hour != 12:
            hour += 12
        if meridian == "am" and hour == 12:
            hour = 0
    elif hour > 23:
        return None

    return hour * 60 + minute


def minutes_to_time_value(minutes):
    """Devuelve HH:MM para inputs/SQL time."""
    minutes = max(0, min(24 * 60, int(minutes)))
    hour = minutes // 60
    minute = minutes % 60
    return f"{hour:02d}:{minute:02d}"


def parse_business_hours_json(raw_value):
    """Lee business_hours_json de la BD y siempre devuelve dict."""
    if not raw_value:
        return {}

    if isinstance(raw_value, dict):
        return raw_value

    try:
        parsed = json.loads(raw_value)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def get_today_business_hours(business, local_dt):
    """
    Obtiene el horario del día actual.
    Formatos soportados: "10:00 - 19:00", "10:00 a 19:00", "10am-7pm", "Cerrado".
    Si el día está vacío, usa 08:00-21:00 para salones sin configuración.
    """
    day_keys = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    business_hours = parse_business_hours_json(business.get("business_hours_json") if hasattr(business, "get") else business["business_hours_json"])
    raw_hours = clean_text(business_hours.get(day_keys[local_dt.weekday()], ""))

    if not raw_hours:
        return {
            "is_open": True,
            "opens_at": "08:00",
            "closes_at": "21:00",
            "label": "08:00 - 21:00",
            "from_config": False,
        }

    if raw_hours.lower() in {"cerrado", "cerrada", "no abre", "closed", "descanso"}:
        return {
            "is_open": False,
            "opens_at": None,
            "closes_at": None,
            "label": raw_hours,
            "from_config": True,
        }

    parts = re.split(r"\s*(?:-|–|—| a | al | hasta )\s*", raw_hours, maxsplit=1, flags=re.IGNORECASE)
    if len(parts) != 2:
        return {
            "is_open": True,
            "opens_at": "08:00",
            "closes_at": "21:00",
            "label": "08:00 - 21:00",
            "from_config": False,
        }

    open_minutes = parse_time_to_minutes(parts[0])
    close_minutes = parse_time_to_minutes(parts[1])

    if open_minutes is None or close_minutes is None or close_minutes <= open_minutes:
        return {
            "is_open": True,
            "opens_at": "08:00",
            "closes_at": "21:00",
            "label": "08:00 - 21:00",
            "from_config": False,
        }

    return {
        "is_open": True,
        "opens_at": minutes_to_time_value(open_minutes),
        "closes_at": minutes_to_time_value(close_minutes),
        "label": f"{minutes_to_time_value(open_minutes)} - {minutes_to_time_value(close_minutes)}",
        "from_config": True,
    }


def parse_business_hours_entry(raw_hours):
    """Convierte un horario guardado en controles para el formulario."""
    raw_hours = clean_text(raw_hours)

    if raw_hours.lower() in {"cerrado", "cerrada", "no abre", "closed", "descanso"}:
        return {"closed": True, "open": "", "close": "", "label": raw_hours or "Cerrado"}

    parts = re.split(r"\s*(?:-|–|—| a | al | hasta )\s*", raw_hours, maxsplit=1, flags=re.IGNORECASE)
    if len(parts) == 2:
        open_minutes = parse_time_to_minutes(parts[0])
        close_minutes = parse_time_to_minutes(parts[1])

        if open_minutes is not None and close_minutes is not None and close_minutes > open_minutes:
            return {
                "closed": False,
                "open": minutes_to_time_value(open_minutes),
                "close": minutes_to_time_value(close_minutes),
                "label": f"{minutes_to_time_value(open_minutes)} - {minutes_to_time_value(close_minutes)}",
            }

    return {"closed": False, "open": "", "close": "", "label": raw_hours}


def normalize_business_hours_from_form(form):
    """Lee switches + horas de configuración y devuelve JSON compatible."""
    day_keys = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    normalized = {}

    for day_key in day_keys:
        is_closed = form.get(f"{day_key}_closed") == "on"
        open_value = clean_text(form.get(f"{day_key}_open"), 8)
        close_value = clean_text(form.get(f"{day_key}_close"), 8)

        if is_closed:
            normalized[day_key] = "Cerrado"
            continue

        if not open_value and not close_value:
            normalized[day_key] = ""
            continue

        if not open_value or not close_value:
            raise ValueError("Completa hora de apertura y cierre, o marca el día como cerrado.")

        open_minutes = parse_time_to_minutes(open_value)
        close_minutes = parse_time_to_minutes(close_value)

        if open_minutes is None or close_minutes is None:
            raise ValueError("Revisa el formato de los horarios.")

        if close_minutes <= open_minutes:
            raise ValueError("La hora de cierre debe ser mayor que la hora de apertura.")

        normalized[day_key] = f"{minutes_to_time_value(open_minutes)} - {minutes_to_time_value(close_minutes)}"

    return normalized


def build_business_hours_controls(business_hours):
    """Prepara los horarios para renderizar switches y time inputs."""
    return {
        day_key: parse_business_hours_entry((business_hours or {}).get(day_key, ""))
        for day_key in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    }


def clean_hex_color(value, default):
    """
    Acepta solo colores HEX de 6 dígitos (#rrggbb).
    Rechaza cualquier otra cadena para evitar CSS inválido o inyección visual.
    """
    color = clean_text(value)
    if re.fullmatch(r"#[0-9a-fA-F]{6}", color):
        return color.lower()
    return default


def clean_optional_id(value):
    """
    Convierte IDs opcionales de formularios a int o None.
    Strings vacíos y ceros se devuelven como None.
    """
    if value in (None, ""):
        return None

    try:
        number = int(value)
    except (TypeError, ValueError):
        return None

    return number if number > 0 else None


def generate_nails_join_code(cur):
    """Genera un código corto único para invitar técnicas a un salón."""
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    for _ in range(20):
        code = "".join(secrets.choice(alphabet) for _ in range(8))
        cur.execute(
            "SELECT id FROM nails_businesses WHERE join_code = %s LIMIT 1",
            (code,),
        )
        if not cur.fetchone():
            return code
    return uuid.uuid4().hex[:10].upper()


def row_belongs_to_business(cur, table, row_id, business_id, active_only=True):
    """
    Valida que un ID recibido por POST pertenezca al salón logueado.
    Evita que un usuario malicioso manipule IDs de otros negocios.
    """
    if not row_id:
        return False

    active_sql = "AND is_active = TRUE" if active_only else ""
    cur.execute(
        f"""
        SELECT id
        FROM {table}
        WHERE id = %s
          AND business_id = %s
          {active_sql}
        LIMIT 1
        """,
        (row_id, business_id),
    )
    return cur.fetchone() is not None


def find_overlapping_appointment(cur, business_id, start_time_db, end_time_db, timezone, staff_id=None, exclude_appointment_id=None):
    """
    Busca una cita activa que se empalme con el rango solicitado.
    Si no hay técnica asignada, revisa toda la agenda del salón.
    Si hay técnica, revisa esa técnica y las citas todavía sin asignar.
    """
    params = [timezone, timezone, business_id, start_time_db, end_time_db]
    staff_filter = ""

    if staff_id:
        staff_filter = "AND (a.staff_id = %s OR a.staff_id IS NULL)"
        params.append(staff_id)

    exclude_filter = ""
    if exclude_appointment_id:
        exclude_filter = "AND a.id != %s"
        params.append(exclude_appointment_id)

    cur.execute(
        f"""
        SELECT
            a.id,
            a.title,
            a.status,
            a.staff_id,
            c.name AS client_name,
            st.name AS staff_name,
            TO_CHAR(a.start_time AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS start_time_label,
            TO_CHAR(a.end_time AT TIME ZONE %s, 'HH24:MI') AS end_time_label
        FROM nails_appointments a
        LEFT JOIN nails_clients c ON c.id = a.client_id
        LEFT JOIN nails_staff st ON st.id = a.staff_id
        WHERE a.business_id = %s
          AND a.status NOT IN ('cancelada', 'no_asistio')
          AND %s < a.end_time
          AND %s > a.start_time
          {staff_filter}
          {exclude_filter}
        ORDER BY a.start_time ASC
        LIMIT 1
        """,
        params,
    )
    return cur.fetchone()


def build_overlap_message(overlap, staff_id=None):
    staff_label = overlap["staff_name"] or "sin técnica asignada"
    client_label = overlap["client_name"] or "Cliente General"

    if staff_id and overlap["staff_id"]:
        return (
            f"Ojo: esta técnica ya tiene una cita programada de "
            f"{overlap['start_time_label']} a {overlap['end_time_label']} "
            f"con {client_label}. Se empalman los horarios."
        )

    if staff_id:
        return (
            f"Ojo: en ese horario ya existe una cita sin técnica asignada, de "
            f"{overlap['start_time_label']} a {overlap['end_time_label']} "
            f"con {client_label}. Revisa la agenda antes de asignar otra cita."
        )

    return (
        f"Ojo: a esta hora ya tienes una cita programada de "
        f"{overlap['start_time_label']} a {overlap['end_time_label']} "
        f"con {client_label} ({staff_label}). Asigna una técnica libre o cambia el horario."
    )


def get_public_available_slots(cur, business, service_ids, appointment_date, extra_ids=None):
    """
    Calcula horarios libres para agendar desde el catálogo público.
    Como el catálogo no asigna técnica, bloquea cualquier cita activa del salón.
    """
    if not isinstance(service_ids, (list, tuple)):
        service_ids = [service_ids]
    service_ids = [clean_optional_id(value) for value in service_ids]
    service_ids = [value for value in service_ids if value]
    unique_service_ids = list(dict.fromkeys(service_ids))
    extra_ids = extra_ids or []
    if not isinstance(extra_ids, (list, tuple)):
        extra_ids = [extra_ids]
    extra_ids = [clean_optional_id(value) for value in extra_ids]
    extra_ids = [value for value in extra_ids if value]
    unique_extra_ids = list(dict.fromkeys(extra_ids))
    appointment_date_value = parse_date_value(appointment_date)

    if not unique_service_ids or not appointment_date_value:
        return [], [], [], "Selecciona al menos un servicio y una fecha válida."

    cur.execute(
        """
        SELECT id, name, base_price, duration_minutes
        FROM nails_services
        WHERE id = ANY(%s)
          AND business_id = %s
          AND is_active = TRUE
          AND is_public = TRUE
        ORDER BY array_position(%s, id)
        """,
        (unique_service_ids, business["id"], unique_service_ids),
    )
    service_rows = cur.fetchall()

    if len(service_rows) != len(unique_service_ids):
        return [], [], [], "Uno o más servicios seleccionados ya no están disponibles."

    service_map = {service["id"]: service for service in service_rows}
    services = [service_map[service_id] for service_id in service_ids]
    extras = []
    if unique_extra_ids:
        cur.execute(
            """
            SELECT id, name, price, duration_minutes
            FROM nails_extras
            WHERE id = ANY(%s)
              AND business_id = %s
              AND is_active = TRUE
            ORDER BY array_position(%s, id)
            """,
            (unique_extra_ids, business["id"], unique_extra_ids),
        )
        extra_rows = cur.fetchall()
        if len(extra_rows) != len(unique_extra_ids):
            return [], [], [], "Uno o más extras seleccionados ya no están disponibles."
        extra_map = {extra["id"]: extra for extra in extra_rows}
        extras = [extra_map[extra_id] for extra_id in extra_ids]

    total_duration = sum(max(15, int(service["duration_minutes"] or 60)) for service in services)
    total_duration += sum(max(0, int(extra["duration_minutes"] or 0)) for extra in extras)

    business_hours = get_today_business_hours(business, appointment_date_value)
    if not business_hours["is_open"]:
        return [], services, extras, "El salón no abre ese día."

    open_minutes = parse_time_to_minutes(business_hours["opens_at"])
    close_minutes = parse_time_to_minutes(business_hours["closes_at"])

    if open_minutes is None or close_minutes is None or close_minutes - open_minutes < total_duration:
        return [], services, extras, "No hay espacio suficiente para esos servicios en el horario del día."

    business_timezone = business["timezone"] or "America/Monterrey"

    cur.execute(
        """
        WITH bounds AS (
            SELECT
                (%s::date + %s::time) AT TIME ZONE %s AS opens_at,
                (%s::date + %s::time) AT TIME ZONE %s AS closes_at
        ),
        slots AS (
            SELECT generate_series(
                bounds.opens_at,
                bounds.closes_at - (%s || ' minutes')::interval,
                INTERVAL '30 minutes'
            ) AS slot_start
            FROM bounds
        )
        SELECT
            TO_CHAR(slot_start AT TIME ZONE %s, 'HH24:MI') AS value,
            TO_CHAR(slot_start AT TIME ZONE %s, 'FMHH12:MI AM') AS label,
            slot_start AS starts_at,
            slot_start + (%s || ' minutes')::interval AS ends_at
        FROM slots
        WHERE slot_start >= NOW()
          AND NOT EXISTS (
              SELECT 1
              FROM nails_appointments a
              WHERE a.business_id = %s
                AND a.status NOT IN ('cancelada', 'no_asistio')
                AND slot_start < a.end_time
                AND slot_start + (%s || ' minutes')::interval > a.start_time
          )
        ORDER BY slot_start ASC
        """,
        (
            appointment_date_value,
            business_hours["opens_at"],
            business_timezone,
            appointment_date_value,
            business_hours["closes_at"],
            business_timezone,
            total_duration,
            business_timezone,
            business_timezone,
            total_duration,
            business["id"],
            total_duration,
        ),
    )
    slots = [
        {
            "value": row["value"],
            "label": row["label"],
        }
        for row in cur.fetchall()
    ]

    if not slots:
        return slots, services, extras, "No hay horarios disponibles para ese día."

    return slots, services, extras, ""


def get_sale_status_from_amounts(total, paid_amount):
    total = float(total or 0)
    paid_amount = float(paid_amount or 0)
    balance_due = max(total - paid_amount, 0)

    if total <= 0 or balance_due <= 0:
        return "pagada"
    if paid_amount > 0:
        return "anticipo"
    return "pendiente"


def sync_sale_for_appointment(
    cur,
    business_id,
    appointment_id,
    client_id,
    staff_id,
    detail_items,
    subtotal,
    initial_paid_amount=0,
    notes="",
):
    """
    Crea o sincroniza la venta ligada a una cita.
    Si la venta ya existe, conserva lo pagado y recalcula saldo/estado.
    """
    subtotal = round(float(subtotal or 0), 2)
    initial_paid_amount = round(min(parse_positive_float(initial_paid_amount), subtotal), 2)

    cur.execute(
        """
        SELECT *
        FROM nails_sales
        WHERE business_id = %s AND appointment_id = %s
        ORDER BY id ASC
        LIMIT 1
        """,
        (business_id, appointment_id),
    )
    sale = cur.fetchone()

    if sale:
        sale_id = sale["id"]
        paid_amount = round(float(sale["paid_amount"] or 0), 2)
        balance_due = max(subtotal - paid_amount, 0)
        status = sale["status"] if sale["status"] == "cancelada" else get_sale_status_from_amounts(subtotal, paid_amount)

        cur.execute(
            """
            UPDATE nails_sales
            SET client_id = %s,
                staff_id = %s,
                subtotal = %s,
                tax_amount = 0,
                total = %s,
                balance_due = %s,
                status = %s,
                notes = COALESCE(NULLIF(%s, ''), notes),
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (
                client_id if client_id else None,
                staff_id if staff_id else None,
                subtotal,
                subtotal,
                balance_due,
                status,
                notes,
                sale_id,
                business_id,
            ),
        )
    else:
        sale_status = get_sale_status_from_amounts(subtotal, initial_paid_amount)
        balance_due = max(subtotal - initial_paid_amount, 0)

        cur.execute(
            """
            INSERT INTO nails_sales (
                business_id, client_id, appointment_id, staff_id,
                subtotal, discount_amount, discount_percentage, tax_amount,
                total, paid_amount, balance_due,
                payment_method, status, notes
            )
            VALUES (%s, %s, %s, %s, %s, 0, 0, 0, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                business_id,
                client_id if client_id else None,
                appointment_id,
                staff_id if staff_id else None,
                subtotal,
                subtotal,
                initial_paid_amount,
                balance_due,
                "efectivo" if initial_paid_amount > 0 else None,
                sale_status,
                notes,
            ),
        )
        sale_id = cur.fetchone()["id"]
        sale_number = f"N-{sale_id:06d}"

        cur.execute(
            "UPDATE nails_sales SET sale_number = %s WHERE id = %s",
            (sale_number, sale_id),
        )

        if initial_paid_amount > 0:
            cur.execute(
                """
                INSERT INTO nails_payments (
                    sale_id, amount, payment_method, payment_type, notes
                )
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    sale_id,
                    initial_paid_amount,
                    "efectivo",
                    "anticipo" if balance_due > 0 else "pago",
                    "Anticipo registrado al crear cita",
                ),
            )

    cur.execute("DELETE FROM nails_sale_details WHERE sale_id = %s", (sale_id,))

    for item in detail_items:
        cur.execute(
            """
            INSERT INTO nails_sale_details (
                sale_id, item_type, item_id, name,
                description, quantity, unit_price, total
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                sale_id,
                item["item_type"],
                item["item_id"],
                item["name"],
                item.get("description"),
                item.get("quantity", 1),
                item["unit_price"],
                item["total"],
            ),
        )

    return sale_id


def user_is_nails_owner(cur, user_id, business) -> bool:
    """
    Valida si la sesión pertenece a la dueña del salón.
    Acepta tanto el owner directo del negocio como un registro staff con role='owner'.
    """
    if not business or not user_id:
        return False

    if int(business["user_id"]) == int(user_id):
        return True

    cur.execute(
        """
        SELECT id
        FROM nails_staff
        WHERE business_id = %s
          AND user_id = %s
          AND role = 'owner'
          AND is_active = TRUE
        LIMIT 1
        """,
        (business["id"], user_id),
    )
    return cur.fetchone() is not None


def get_current_nails_role(cur, user_id, business):
    """Devuelve el rol Nails del usuario actual para mostrarlo en la interfaz."""
    if not business or not user_id:
        return "Usuario Nails"

    role = get_current_nails_role_key(cur, user_id, business)
    if role:
        return NAILS_ROLE_LABELS.get(role, role.title())

    if int(business["user_id"]) == int(user_id):
        return "Jefa"

    return "Usuario Nails"


NAILS_ROLE_LABELS = {
    "owner": "Jefa",
    "admin": "Administración",
    "staff": "Técnica",
    "reception": "Recepción",
}

NAILS_ROLE_DESCRIPTIONS = {
    "owner": "Control total del salón: configuración, personal, citas, ventas, reportes y acciones sensibles.",
    "admin": "Apoya la operación diaria: agenda, ventas, clientas, servicios y reportes. No administra jefas.",
    "staff": "Técnica del salón: aparece para asignar citas y ventas; enfocada en servicios y seguimiento de clientas.",
    "reception": "Recepción: apoyo en agenda, clientas y cobros sin administrar configuración sensible.",
}

NAILS_STAFF_ROLES = set(NAILS_ROLE_LABELS.keys())

NAILS_ROLE_PERMISSIONS = {
    "owner": {
        "dashboard", "agenda", "clientes", "servicios", "ventas",
        "gastos", "reportes", "galeria", "personal", "configuracion",
    },
    "admin": {"dashboard", "agenda", "clientes", "servicios", "ventas", "gastos", "reportes", "galeria"},
    "staff": {"dashboard", "agenda", "clientes", "ventas"},
    "reception": {"dashboard", "agenda", "clientes", "ventas"},
}

NAILS_ENDPOINT_SECTIONS = {
    "index": "dashboard",
    "api_clientes": "clientes",
    "dashboard": "dashboard",
    "agenda": "agenda",
    "cambiar_estado_cita": "agenda",
    "editar_cita": "agenda",
    "eliminar_cita": "agenda",
    "clientes": "clientes",
    "servicios": "servicios",
    "editar_servicio": "servicios",
    "eliminar_servicio": "servicios",
    "editar_extra": "servicios",
    "eliminar_extra": "servicios",
    "ventas": "ventas",
    "editar_venta": "ventas",
    "anular_venta": "ventas",
    "ticket": "ventas",
    "gastos": "gastos",
    "editar_gasto": "gastos",
    "cancelar_gasto": "gastos",
    "reportes": "reportes",
    "galeria": "galeria",
    "eliminar_imagen_galeria": "galeria",
    "toggle_public_galeria": "galeria",
    "personal": "personal",
    "editar_personal": "personal",
    "desactivar_personal": "personal",
    "configuracion": "configuracion",
    "upload_r2_nails": "galeria",
}

NAILS_PUBLIC_ENDPOINTS = {
    "cliente_acceso",
    "catalogo_publico",
    "catalogo_horarios_disponibles",
}

NAILS_UNRESTRICTED_ENDPOINTS = {
    "onboarding",
}


def get_current_nails_role_key(cur, user_id, business):
    """Devuelve la llave de rol interna para aplicar permisos."""
    if not business or not user_id:
        return None

    if int(business["user_id"]) == int(user_id):
        return "owner"

    cur.execute(
        """
        SELECT role
        FROM nails_staff
        WHERE business_id = %s
          AND user_id = %s
          AND is_active = TRUE
        ORDER BY
            CASE role WHEN 'owner' THEN 1 WHEN 'admin' THEN 2 WHEN 'reception' THEN 3 ELSE 4 END,
            id DESC
        LIMIT 1
        """,
        (business["id"], user_id),
    )
    staff = cur.fetchone()
    if staff and staff["role"] in NAILS_STAFF_ROLES:
        return staff["role"]
    return None


def user_can_access_nails_section(cur, user_id, business, section):
    role = get_current_nails_role_key(cur, user_id, business)
    return section in NAILS_ROLE_PERMISSIONS.get(role, set())


@nails_bp.context_processor
def inject_nails_role_context():
    csrf_context = nails_csrf_context()
    if not request.endpoint or not request.endpoint.startswith("nails."):
        return csrf_context

    user_id = session.get("user_id")
    if not user_id:
        return csrf_context

    business = get_user_nails_business(user_id)
    if not business:
        return csrf_context

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        role_key = get_current_nails_role_key(cur, user_id, business)
        permissions = NAILS_ROLE_PERMISSIONS.get(role_key, set())
        return {
            **csrf_context,
            "nails_role_key": role_key,
            "nails_role_label": NAILS_ROLE_LABELS.get(role_key, "Usuario Nails"),
            "nails_permissions": permissions,
        }
    finally:
        cur.close()
        conn.close()


@nails_bp.before_request
def enforce_nails_role_permissions():
    if not request.endpoint or not request.endpoint.startswith("nails."):
        return None

    endpoint_name = request.endpoint.split(".", 1)[1]
    if endpoint_name in NAILS_PUBLIC_ENDPOINTS or endpoint_name in NAILS_UNRESTRICTED_ENDPOINTS:
        if request.method == "POST" and not validate_nails_csrf():
            flash("Tu sesión de seguridad expiró. Recarga la página e intenta de nuevo.", "warning")
            return nails_csrf_error_response()
        return None

    section = NAILS_ENDPOINT_SECTIONS.get(endpoint_name)
    if not section:
        return None

    user_id = session.get("user_id")
    if not user_id:
        if endpoint_name.startswith("api_") or request.accept_mimetypes.best == "application/json":
            return jsonify({"success": False, "error": "No autorizado"}), 401
        return redirect(url_for("auth.login"))

    if request.method == "POST" and not validate_nails_csrf():
        flash("Tu sesión de seguridad expiró. Recarga la página e intenta de nuevo.", "warning")
        return nails_csrf_error_response()

    business = get_user_nails_business(user_id)
    if not business:
        return None

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        if user_can_access_nails_section(cur, user_id, business, section):
            return None
    finally:
        cur.close()
        conn.close()

    flash("Tu rol no tiene permiso para entrar a esa sección.", "warning")
    return redirect(url_for("nails.dashboard"))


def ensure_nails_appointment_services_table(cur):
    """
    Crea la tabla de servicios por cita si todavía no existe.
    Es idempotente y protege instalaciones existentes antes de correr db.py.
    """
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS nails_appointment_services (
            id SERIAL PRIMARY KEY,
            appointment_id INTEGER NOT NULL REFERENCES nails_appointments(id) ON DELETE CASCADE,
            service_id INTEGER REFERENCES nails_services(id) ON DELETE SET NULL,

            name TEXT NOT NULL,
            price NUMERIC(10,2) DEFAULT 0,
            duration_minutes INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0,

            created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_nails_appointment_services_appointment_id
        ON nails_appointment_services(appointment_id)
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_nails_appointment_services_service_id
        ON nails_appointment_services(service_id)
        """
    )


def ensure_nails_expenses_table(cur):
    """Crea la tabla de gastos de Nails si la migración principal aún no corrió."""
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS nails_expenses (
            id SERIAL PRIMARY KEY,
            business_id INTEGER NOT NULL REFERENCES nails_businesses(id) ON DELETE CASCADE,

            title TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT 'otros',

            amount NUMERIC(10,2) NOT NULL DEFAULT 0,

            expense_date DATE NOT NULL DEFAULT CURRENT_DATE,
            payment_method TEXT DEFAULT 'efectivo',

            is_recurring BOOLEAN DEFAULT FALSE,
            recurring_day INTEGER,
            recurring_frequency TEXT,

            notes TEXT,

            status TEXT DEFAULT 'activo',

            created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'chk_nails_expenses_category') THEN
                ALTER TABLE nails_expenses
                ADD CONSTRAINT chk_nails_expenses_category
                CHECK (category IN (
                    'materiales', 'renta', 'servicios', 'sueldos', 'comisiones',
                    'publicidad', 'mantenimiento', 'capacitacion', 'otros'
                ));
            END IF;
        END $$
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'chk_nails_expenses_payment_method') THEN
                ALTER TABLE nails_expenses
                ADD CONSTRAINT chk_nails_expenses_payment_method
                CHECK (payment_method IN ('efectivo', 'transferencia', 'tarjeta', 'mixto', 'otro'));
            END IF;
        END $$
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'chk_nails_expenses_recurring_frequency') THEN
                ALTER TABLE nails_expenses
                ADD CONSTRAINT chk_nails_expenses_recurring_frequency
                CHECK (
                    recurring_frequency IS NULL
                    OR recurring_frequency IN ('semanal', 'quincenal', 'mensual', 'bimestral', 'anual')
                );
            END IF;
        END $$
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'chk_nails_expenses_status') THEN
                ALTER TABLE nails_expenses
                ADD CONSTRAINT chk_nails_expenses_status
                CHECK (status IN ('activo', 'cancelado'));
            END IF;
        END $$
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'chk_nails_expenses_amount_positive') THEN
                ALTER TABLE nails_expenses
                ADD CONSTRAINT chk_nails_expenses_amount_positive
                CHECK (amount >= 0);
            END IF;
        END $$
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'chk_nails_expenses_recurring_day') THEN
                ALTER TABLE nails_expenses
                ADD CONSTRAINT chk_nails_expenses_recurring_day
                CHECK (recurring_day IS NULL OR recurring_day BETWEEN 1 AND 31);
            END IF;
        END $$
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_nails_expenses_business_id ON nails_expenses(business_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_nails_expenses_category ON nails_expenses(category)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_nails_expenses_expense_date ON nails_expenses(expense_date DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_nails_expenses_business_date ON nails_expenses(business_id, expense_date DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_nails_expenses_business_category ON nails_expenses(business_id, category)")
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_nails_expenses_recurring
        ON nails_expenses(business_id, is_recurring)
        WHERE is_recurring = TRUE
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_nails_expenses_status ON nails_expenses(status)")


def image_url_to_data_uri(image_url):
    """
    Descarga una imagen remota y la convierte a Data URI base64.
    Se usa para incrustar el logo en el ticket sin depender de una URL
    externa al momento de impresión/visualización offline.
    Devuelve None si la descarga falla (el caller decide el fallback).
    """
    if not image_url:
        return None

    try:
        # timeout=(connect, read) para no colgar el hilo en descargas lentas.
        response = requests.get(image_url, timeout=(5, 8))
        response.raise_for_status()

        content_type = response.headers.get("Content-Type", "").split(";")[0].strip()
        if not content_type.startswith("image/"):
            content_type = "image/png"

        encoded = base64.b64encode(response.content).decode("ascii")
        return f"data:{content_type};base64,{encoded}"

    except Exception as e:
        current_app.logger.warning(
            f"NAILS_LOGO_DATA_URI_WARNING: No se pudo preparar logo para ticket - {e}"
        )
        return None


def generate_slug(text):
    """
    Genera un slug URL-amigable a partir de texto arbitrario.
    Ejemplo: 'Salón Dé Uñas' → 'salon-de-unas'
    BUG CORREGIDO: faltaba el reemplazo de 'ü' → 'u'.
    """
    text = text.lower().strip()
    # Conservar solo letras (con acentos/ñ/ü), dígitos, espacios y guiones
    text = re.sub(r"[^a-z0-9áéíóúñü\s-]", "", text)
    # Espacios → guiones
    text = re.sub(r"\s+", "-", text)
    # Reemplazar caracteres especiales del español
    text = (
        text
        .replace("ñ", "n")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ü", "u")   # ← BUG CORREGIDO: antes faltaba este reemplazo
    )
    # Colapsar guiones múltiples y limpiar extremos
    text = re.sub(r"-+", "-", text)
    return text.strip("-")


def require_login():
    """Devuelve True si hay una sesión activa de usuario."""
    return "user_id" in session


def get_user_nails_business(user_id):
    """
    Obtiene el negocio Nails activo del usuario logueado.
    Devuelve None si el usuario no tiene ningún salón registrado.
    """
    conn = get_db_connection()
    cur  = conn.cursor()

    cur.execute(
        """
        SELECT b.*
        FROM nails_businesses b
        WHERE b.user_id = %s
          AND b.is_active = TRUE

        UNION ALL

        SELECT b.*
        FROM nails_businesses b
        INNER JOIN nails_staff st ON st.business_id = b.id
        WHERE st.user_id = %s
          AND st.is_active = TRUE
          AND b.is_active = TRUE

        ORDER BY id DESC
        LIMIT 1
        """,
        (user_id, user_id),
    )

    business = cur.fetchone()

    cur.close()
    conn.close()

    return business


# =========================================================
# RUTA RAÍZ
# Redirige al dashboard; punto de entrada del módulo.
# =========================================================

@nails_bp.route("/")
def index():
    return redirect(url_for("nails.dashboard"))


# =========================================================
# ACCESO CLIENTA
# Pantalla pública para que una clienta entre al catálogo
# de su salón con el código compartido.
# =========================================================

@nails_bp.route("/cliente", methods=["GET", "POST"])
@nails_bp.route("/cliente/<codigo>", methods=["GET"])
def cliente_acceso(codigo=None):
    code = clean_text(codigo or request.values.get("codigo"), 80).upper()

    def find_client_catalog_slug(raw_code):
        normalized_slug = generate_slug(raw_code)
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                SELECT slug
                FROM nails_businesses
                WHERE is_active = TRUE
                  AND (
                    UPPER(join_code) = %s
                    OR slug = %s
                  )
                LIMIT 1
                """,
                (raw_code, normalized_slug),
            )
            business = cur.fetchone()
            return business["slug"] if business else None
        finally:
            cur.close()
            conn.close()

    if request.method == "GET" and code:
        slug = find_client_catalog_slug(code)
        if slug:
            return redirect(url_for("nails.catalogo_publico", slug=slug))
        flash("No encontramos un salón activo con ese código.", "warning")

    if request.method == "POST":
        code = clean_text(request.form.get("codigo"), 80).upper()

        if not code:
            flash("Escribe el código del salón para entrar como clienta.", "warning")
            return render_template("nails/cliente_acceso.html", code=code)

        slug = find_client_catalog_slug(code)
        if not slug:
            flash("No encontramos un salón activo con ese código.", "warning")
            return render_template("nails/cliente_acceso.html", code=code)

        return redirect(url_for("nails.catalogo_publico", slug=slug))

    return render_template("nails/cliente_acceso.html", code=code)


# =========================================================
# API INTERNA: BUSCADOR DE CLIENTAS
# Usado por el autocomplete de Agenda. Solo devuelve clientas
# activas del salón del usuario en sesión.
# =========================================================

@nails_bp.route("/api/clientes")
def api_clientes():
    if not require_login():
        return jsonify({"success": False, "error": "No autorizado"}), 401

    user_id = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return jsonify({"success": True, "clientes": []})

    q = clean_text(request.args.get("q"), 80)
    if len(q) < 2 or q.lower() == "cliente general":
        return jsonify({"success": True, "clientes": []})

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        like = f"%{q}%"
        cur.execute("""
            SELECT id, name, phone, instagram
            FROM nails_clients
            WHERE business_id = %s
              AND is_active = TRUE
              AND (
                TRANSLATE(LOWER(name), 'áéíóú', 'aeiou') ILIKE TRANSLATE(LOWER(%s), 'áéíóú', 'aeiou')
                OR COALESCE(phone, '') ILIKE %s
                OR COALESCE(instagram, '') ILIKE %s
              )
            ORDER BY LOWER(name) ASC
            LIMIT 8
        """, (business["id"], like, like, like))

        return jsonify({
            "success": True,
            "clientes": [dict(row) for row in cur.fetchall()]
        })

    finally:
        cur.close()
        conn.close()


# =========================================================
# DASHBOARD / INICIO
# Muestra la operación del día: KPIs, agenda, próxima cita,
# saldos pendientes, huecos disponibles y servicios más reservados.
# =========================================================

@nails_bp.route("/dashboard")
def dashboard():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        business_timezone = business["timezone"] or "America/Monterrey"

        # ── Ventas de hoy ──────────────────────────────────
        cur.execute(
            """
            SELECT
                COALESCE(SUM(total), 0)  AS total_today,
                COUNT(*)                 AS sales_count_today
            FROM nails_sales
            WHERE business_id = %s
              AND status != 'cancelada'
              AND DATE(created_at AT TIME ZONE %s) = DATE(NOW() AT TIME ZONE %s)
            """,
            (business["id"], business_timezone, business_timezone),
        )
        sales_today = cur.fetchone()

        # ── Citas de hoy (excluye canceladas y no-shows) ───
        cur.execute(
            """
            SELECT
                COUNT(*) AS appointments_today,
                COUNT(*) FILTER (WHERE status = 'confirmada') AS confirmed_today,
                COUNT(*) FILTER (
                    WHERE start_time >= NOW()
                      AND start_time <= NOW() + INTERVAL '30 minutes'
                      AND status IN ('pendiente', 'confirmada')
                ) AS arriving_soon
            FROM nails_appointments
            WHERE business_id = %s
              AND DATE(start_time AT TIME ZONE %s) = DATE(NOW() AT TIME ZONE %s)
              AND status NOT IN ('cancelada', 'no_asistio')
            """,
            (business["id"], business_timezone, business_timezone),
        )
        appointments_today = cur.fetchone()

        # ── Total de clientas activas ──────────────────────
        cur.execute(
            """
            SELECT COUNT(*) AS total_clients
            FROM nails_clients
            WHERE business_id = %s
              AND is_active = TRUE
            """,
            (business["id"],),
        )
        clients_count = cur.fetchone()

        # ── Saldo pendiente (ventas sin liquidar) ──────────
        cur.execute(
            """
            SELECT COALESCE(SUM(balance_due), 0) AS pending_balance
            FROM nails_sales
            WHERE business_id = %s
              AND status IN ('pendiente', 'anticipo')
            """,
            (business["id"],),
        )
        pending_balance = cur.fetchone()

        # ── Ventas pendientes de pago ─────────────────────
        cur.execute(
            """
            SELECT COUNT(*) AS pending_payments_count
            FROM nails_sales
            WHERE business_id = %s
              AND status IN ('pendiente', 'anticipo')
              AND COALESCE(balance_due, 0) > 0
            """,
            (business["id"],),
        )
        pending_payments_count = cur.fetchone()

        # ── Agenda de hoy ─────────────────────────────────
        cur.execute(
            """
            SELECT
                a.id,
                a.title,
                a.status,
                a.estimated_total,
                a.deposit_amount,
                a.reminder_sent,
                a.start_time,
                a.end_time,
                c.name  AS client_name,
                c.phone AS client_phone,
                s.name  AS service_name,
                st.name AS staff_name,
                TO_CHAR(a.start_time AT TIME ZONE %s, 'HH12:MI AM') AS start_time_label,
                TO_CHAR(a.end_time AT TIME ZONE %s, 'HH12:MI AM') AS end_time_label,
                GREATEST(15, ROUND(EXTRACT(EPOCH FROM (a.end_time - a.start_time)) / 60)::INT) AS duration_minutes
            FROM nails_appointments a
            LEFT JOIN nails_clients  c  ON c.id  = a.client_id
            LEFT JOIN nails_services s  ON s.id  = a.service_id
            LEFT JOIN nails_staff    st ON st.id = a.staff_id
            WHERE a.business_id = %s
              AND DATE(a.start_time AT TIME ZONE %s) = DATE(NOW() AT TIME ZONE %s)
              AND a.status NOT IN ('cancelada', 'no_asistio')
            ORDER BY a.start_time ASC
            LIMIT 8
            """,
            (business_timezone, business_timezone, business["id"], business_timezone, business_timezone),
        )
        today_appointments = cur.fetchall()

        # ── Próximas 5 citas (pendientes o confirmadas) ────
        cur.execute(
            """
            SELECT
                a.*,
                c.name  AS client_name,
                c.phone AS client_phone,
                s.name  AS service_name,
                st.name AS staff_name,
                TO_CHAR(a.start_time AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS start_time_formatted
            FROM nails_appointments a
            LEFT JOIN nails_clients  c  ON c.id  = a.client_id
            LEFT JOIN nails_services s  ON s.id  = a.service_id
            LEFT JOIN nails_staff    st ON st.id = a.staff_id
            WHERE a.business_id = %s
              AND a.start_time  >= NOW()
              AND a.status IN ('pendiente', 'confirmada')
            ORDER BY a.start_time ASC
            LIMIT 5
            """,
            (business_timezone, business["id"]),
        )
        upcoming_appointments = cur.fetchall()

        next_appointment = upcoming_appointments[0] if upcoming_appointments else None

        # ── Últimas 5 ventas ───────────────────────────────
        cur.execute(
            """
            SELECT
                v.*,
                c.name AS client_name,
                TO_CHAR(v.created_at AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS created_at_formatted
            FROM nails_sales v
            LEFT JOIN nails_clients c ON c.id = v.client_id
            WHERE v.business_id = %s
            ORDER BY v.created_at DESC
            LIMIT 5
            """,
            (business_timezone, business["id"]),
        )
        recent_sales = cur.fetchall()

        # ── Saldos pendientes destacados ───────────────────
        cur.execute(
            """
            SELECT
                v.id,
                v.sale_number,
                v.total,
                v.balance_due,
                v.status,
                c.name AS client_name,
                s.name AS service_name
            FROM nails_sales v
            LEFT JOIN nails_clients c ON c.id = v.client_id
            LEFT JOIN nails_appointments a ON a.id = v.appointment_id
            LEFT JOIN nails_services s ON s.id = a.service_id
            WHERE v.business_id = %s
              AND v.status IN ('pendiente', 'anticipo')
              AND COALESCE(v.balance_due, 0) > 0
            ORDER BY v.created_at DESC
            LIMIT 4
            """,
            (business["id"],),
        )
        pending_sales = cur.fetchall()

        # ── Clientes que visitan hoy ───────────────────────
        cur.execute(
            """
            SELECT DISTINCT ON (COALESCE(c.id, a.id))
                c.id,
                COALESCE(c.name, 'Cliente General') AS name,
                c.phone,
                c.birthday,
                c.allergies_notes,
                c.total_visits,
                s.name AS service_name,
                TO_CHAR(a.start_time AT TIME ZONE %s, 'HH12:MI AM') AS appointment_time
            FROM nails_appointments a
            LEFT JOIN nails_clients c ON c.id = a.client_id
            LEFT JOIN nails_services s ON s.id = a.service_id
            WHERE a.business_id = %s
              AND DATE(a.start_time AT TIME ZONE %s) = DATE(NOW() AT TIME ZONE %s)
              AND a.status NOT IN ('cancelada', 'no_asistio')
            ORDER BY COALESCE(c.id, a.id), a.start_time ASC
            LIMIT 4
            """,
            (business_timezone, business["id"], business_timezone, business_timezone),
        )
        today_clients = cur.fetchall()

        now_local = datetime.utcnow()
        try:
            from zoneinfo import ZoneInfo
            now_local = datetime.now(ZoneInfo(business_timezone))
        except Exception:
            pass

        today_business_hours = get_today_business_hours(business, now_local)
        available_slots = []

        # ── Huecos disponibles hoy según horario configurado ─
        if today_business_hours["is_open"]:
            cur.execute(
                """
                WITH bounds AS (
                    SELECT
                        ((DATE(NOW() AT TIME ZONE %s) + %s::time) AT TIME ZONE %s) AS opens_at,
                        ((DATE(NOW() AT TIME ZONE %s) + %s::time) AT TIME ZONE %s) AS closes_at
                ),
                slots AS (
                    SELECT generate_series(opens_at, closes_at - INTERVAL '30 minutes', INTERVAL '30 minutes') AS slot_start
                    FROM bounds
                )
                SELECT
                    TO_CHAR(slot_start AT TIME ZONE %s, 'HH12:MI AM') AS time_label,
                    TO_CHAR(slot_start AT TIME ZONE %s, 'YYYY-MM-DD') AS date_value,
                    TO_CHAR(slot_start AT TIME ZONE %s, 'HH24:MI') AS time_value
                FROM slots
                WHERE slot_start >= NOW()
                  AND NOT EXISTS (
                    SELECT 1
                    FROM nails_appointments a
                    WHERE a.business_id = %s
                      AND a.status NOT IN ('cancelada', 'no_asistio')
                      AND slot_start < a.end_time
                      AND slot_start + INTERVAL '30 minutes' > a.start_time
                  )
                ORDER BY slot_start ASC
                LIMIT 5
                """,
                (
                    business_timezone,
                    today_business_hours["opens_at"],
                    business_timezone,
                    business_timezone,
                    today_business_hours["closes_at"],
                    business_timezone,
                    business_timezone,
                    business_timezone,
                    business_timezone,
                    business["id"],
                ),
            )
            available_slots = cur.fetchall()

        # ── Servicios más reservados hoy ───────────────────
        cur.execute(
            """
            WITH services_today AS (
                SELECT a.id AS appointment_id, a.service_id
                FROM nails_appointments a
                WHERE a.business_id = %s
                  AND a.service_id IS NOT NULL
                  AND DATE(a.start_time AT TIME ZONE %s) = DATE(NOW() AT TIME ZONE %s)
                  AND a.status NOT IN ('cancelada', 'no_asistio')

                UNION ALL

                SELECT a.id AS appointment_id, aps.service_id
                FROM nails_appointment_services aps
                INNER JOIN nails_appointments a ON a.id = aps.appointment_id
                WHERE a.business_id = %s
                  AND aps.service_id IS NOT NULL
                  AND DATE(a.start_time AT TIME ZONE %s) = DATE(NOW() AT TIME ZONE %s)
                  AND a.status NOT IN ('cancelada', 'no_asistio')
            )
            SELECT s.name, COUNT(DISTINCT st.appointment_id) AS bookings_count
            FROM services_today st
            INNER JOIN nails_services s ON s.id = st.service_id
            GROUP BY s.id, s.name
            ORDER BY bookings_count DESC, s.name ASC
            LIMIT 3
            """,
            (
                business["id"],
                business_timezone,
                business_timezone,
                business["id"],
                business_timezone,
                business_timezone,
            ),
        )
        top_services_today = cur.fetchall()

        stats = {
            "total_today":         sales_today["total_today"]              or 0,
            "sales_count_today":   sales_today["sales_count_today"]        or 0,
            "appointments_today":  appointments_today["appointments_today"] or 0,
            "confirmed_today":     appointments_today["confirmed_today"]    or 0,
            "arriving_soon":       appointments_today["arriving_soon"]      or 0,
            "total_clients":       clients_count["total_clients"]           or 0,
            "pending_balance":     pending_balance["pending_balance"]       or 0,
            "pending_payments_count": pending_payments_count["pending_payments_count"] or 0,
        }

        return render_template(
            "nails/dashboard.html",
            business=business,
            stats=stats,
            now_local=now_local,
            current_nails_role=get_current_nails_role(cur, user_id, business),
            today_appointments=today_appointments,
            upcoming_appointments=upcoming_appointments,
            next_appointment=next_appointment,
            recent_sales=recent_sales,
            pending_sales=pending_sales,
            today_clients=today_clients,
            available_slots=available_slots,
            today_business_hours=today_business_hours,
            top_services_today=top_services_today,
        )

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error al cargar Inicio: {e}", "danger")
        return redirect(url_for("nails.agenda"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# ONBOARDING
# Formulario de alta del salón. Solo accesible si el usuario
# no tiene ningún negocio Nails registrado todavía.
# Crea el negocio, su slug único y un registro de staff con
# rol 'owner' para el usuario que lo creó.
# =========================================================

@nails_bp.route("/onboarding", methods=["GET", "POST"])
def onboarding():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)
    prefill = session.get("nails_onboarding_prefill") or {}

    if business:
        return redirect(url_for("nails.dashboard"))

    if request.method == "POST":
        name            = clean_text(request.form.get("name"), 120)
        whatsapp        = clean_text(request.form.get("whatsapp"), 40)
        instagram       = clean_text(request.form.get("instagram"), 80)
        address         = clean_text(request.form.get("address"), 240)
        primary_color   = clean_hex_color(request.form.get("primary_color"),   DEFAULT_PRIMARY_COLOR)
        secondary_color = clean_hex_color(request.form.get("secondary_color"), DEFAULT_SECONDARY_COLOR)
        accent_color    = clean_hex_color(request.form.get("accent_color"),    DEFAULT_ACCENT_COLOR)
        catalog_tagline = clean_text(request.form.get("catalog_tagline"), 240)

        if not name:
            flash("El nombre del salón es obligatorio.", "warning")
            return render_template(
                "nails/onboarding.html",
                business=None,
                prefill=prefill,
                business_hours_controls=build_business_hours_controls({}),
            )

        try:
            business_hours_json = normalize_business_hours_from_form(request.form)
        except ValueError as e:
            flash(str(e), "warning")
            return render_template(
                "nails/onboarding.html",
                business=None,
                prefill={
                    "name": name,
                    "whatsapp": whatsapp,
                    "instagram": instagram,
                    "address": address,
                    "catalog_tagline": catalog_tagline,
                },
                business_hours_controls=build_business_hours_controls({}),
            )

        base_slug = generate_slug(name)
        slug      = base_slug

        conn = get_db_connection()
        cur  = conn.cursor()

        try:
            # Garantiza unicidad del slug incrementando un contador
            counter = 1
            while True:
                cur.execute(
                    "SELECT id FROM nails_businesses WHERE slug = %s LIMIT 1",
                    (slug,),
                )
                if not cur.fetchone():
                    break
                slug = f"{base_slug}-{counter}"
                counter += 1

            # Crea el negocio
            join_code = generate_nails_join_code(cur)
            cur.execute(
                """
                INSERT INTO nails_businesses (
                    user_id, name, slug, whatsapp, instagram,
                    address, primary_color, secondary_color, accent_color,
                    catalog_tagline, business_hours_json, join_code
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (user_id, name, slug, whatsapp, instagram,
                 address, primary_color, secondary_color, accent_color,
                 catalog_tagline or 'Uñas que expresan tu estilo, hechas con amor y detalle.',
                 json.dumps(business_hours_json, ensure_ascii=False), join_code),
            )
            business_id = cur.fetchone()["id"]

            # Registra al dueño como primer miembro del staff (rol 'owner')
            cur.execute(
                """
                INSERT INTO nails_staff (business_id, user_id, name, email, role, color)
                SELECT %s, id,
                       COALESCE(username, company_name, email, 'Dueña'),
                       email, 'owner', %s
                FROM usuarios
                WHERE id = %s
                """,
                (business_id, primary_color, user_id),
            )

            # Activa el módulo Nails en el perfil del usuario
            cur.execute(
                "UPDATE usuarios SET active_module = 'nails' WHERE id = %s",
                (user_id,),
            )

            conn.commit()
            session.pop("nails_onboarding_prefill", None)
            flash("Tu salón fue configurado correctamente.", "success")
            return redirect(url_for("nails.dashboard"))

        except Exception as e:
            conn.rollback()
            flash(f"No se pudo crear el salón: {e}", "danger")
            return render_template(
                "nails/onboarding.html",
                business=None,
                prefill=prefill,
                business_hours_controls=build_business_hours_controls({}),
            )

        finally:
            cur.close()
            conn.close()

    if not prefill:
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                SELECT company_name, telefono
                FROM usuarios
                WHERE id = %s
                LIMIT 1
                """,
                (user_id,),
            )
            user = cur.fetchone()
            if user:
                prefill = {
                    "name": user["company_name"] or "",
                    "whatsapp": user["telefono"] or "",
                }
        finally:
            cur.close()
            conn.close()

    default_business_hours = {
        "monday": "09:00 - 18:00",
        "tuesday": "09:00 - 18:00",
        "wednesday": "09:00 - 18:00",
        "thursday": "09:00 - 18:00",
        "friday": "09:00 - 18:00",
        "saturday": "10:00 - 16:00",
        "sunday": "Cerrado",
    }
    return render_template(
        "nails/onboarding.html",
        business=None,
        prefill=prefill,
        business_hours_controls=build_business_hours_controls(default_business_hours),
    )


# =========================================================
# AGENDA
# GET:  Lista las últimas 50 citas del salón con extras incluidos.
# POST: Crea una nueva cita, opcionalmente creando también una
#       clienta nueva ("clienta rápida") si no se eligió una existente.
#       Valida ownership de todos los IDs para evitar manipulación.
# =========================================================

@nails_bp.route("/agenda", methods=["GET", "POST"])
def agenda():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        ensure_nails_appointment_services_table(cur)
        conn.commit()

        if request.method == "POST":
            # ── Leer y sanitizar campos del formulario ─────
            client_id          = clean_optional_id(request.form.get("client_id"))
            quick_client_name  = clean_text(request.form.get("quick_client_name"), 120)
            quick_client_phone = clean_text(request.form.get("quick_client_phone"), 40)
            staff_id           = clean_optional_id(request.form.get("staff_id"))
            selected_service_ids = [
                clean_optional_id(value)
                for value in request.form.getlist("service_ids")
            ]
            selected_service_ids = [value for value in selected_service_ids if value]
            service_id_from_picker = clean_optional_id(request.form.get("service_id"))
            if not selected_service_ids and service_id_from_picker:
                selected_service_ids = [service_id_from_picker]
            service_id         = selected_service_ids[0] if selected_service_ids else None
            appointment_date   = clean_text(request.form.get("appointment_date"))
            start_time         = clean_text(request.form.get("start_time"))
            status             = clean_text(request.form.get("status", "pendiente"))
            deposit_amount_raw = request.form.get("deposit_amount", "0")
            notes              = clean_text(request.form.get("notes"), 1000)
            selected_extras    = request.form.getlist("extras")

            # ── Validaciones básicas ───────────────────────
            if not service_id or not appointment_date or not start_time:
                flash("Servicio, fecha y hora son obligatorios.", "warning")
                return redirect(url_for("nails.agenda"))

            if status not in APPOINTMENT_STATUSES:
                flash("Estado de cita inválido.", "warning")
                return redirect(url_for("nails.agenda"))

            if not parse_date_value(appointment_date) or not re.fullmatch(r"\d{2}:\d{2}", start_time):
                flash("Fecha u hora inválida.", "warning")
                return redirect(url_for("nails.agenda"))

            deposit_amount = parse_positive_float(deposit_amount_raw)

            # ── Validar ownership de IDs foráneos ──────────
            if client_id and not row_belongs_to_business(cur, "nails_clients", client_id, business["id"]):
                flash("La clienta seleccionada no pertenece a este salón.", "warning")
                return redirect(url_for("nails.agenda"))

            if staff_id and not row_belongs_to_business(cur, "nails_staff", staff_id, business["id"]):
                flash("La técnica seleccionada no pertenece a este salón.", "warning")
                return redirect(url_for("nails.agenda"))

            # ── Creación rápida de clienta nueva ──────────
            if not client_id and quick_client_name and not is_general_customer_name(quick_client_name):
                cur.execute(
                    """
                    INSERT INTO nails_clients (business_id, name, phone)
                    VALUES (%s, %s, %s)
                    RETURNING id
                    """,
                    (business["id"], quick_client_name, quick_client_phone),
                )
                client_id = cur.fetchone()["id"]

            # ── Obtener servicios y calcular precio/duración ──
            unique_service_ids = list(dict.fromkeys(selected_service_ids))

            cur.execute(
                """
                SELECT id, name, base_price, duration_minutes
                FROM nails_services
                WHERE id = ANY(%s)
                  AND business_id = %s
                  AND is_active = TRUE
                ORDER BY array_position(%s, id)
                """,
                (unique_service_ids, business["id"], unique_service_ids),
            )
            service_rows = cur.fetchall()

            if len(service_rows) != len(unique_service_ids):
                flash("Uno o más servicios seleccionados no existen.", "warning")
                return redirect(url_for("nails.agenda"))

            service_map = {service["id"]: service for service in service_rows}
            services_to_insert = []
            total_duration = 0
            estimated_total = 0

            for index, selected_service_id in enumerate(selected_service_ids):
                service = service_map[selected_service_id]
                service_price = float(service["base_price"] or 0)
                service_duration = int(service["duration_minutes"] or 60)
                estimated_total += service_price
                total_duration += service_duration
                services_to_insert.append({
                    "id": service["id"],
                    "name": service["name"],
                    "price": service_price,
                    "duration_minutes": service_duration,
                    "sort_order": index,
                })

            extras_to_insert = []

            # ── Acumular extras válidos ────────────────────
            for extra_id in selected_extras:
                cur.execute(
                    """
                    SELECT id, name, price, duration_minutes
                    FROM nails_extras
                    WHERE id = %s
                      AND business_id = %s
                      AND is_active = TRUE
                    LIMIT 1
                    """,
                    (extra_id, business["id"]),
                )
                extra = cur.fetchone()

                if extra:
                    extra_price    = float(extra["price"]            or 0)
                    extra_duration = int(extra["duration_minutes"]   or 0)
                    estimated_total += extra_price
                    total_duration  += extra_duration
                    extras_to_insert.append({
                        "id":               extra["id"],
                        "name":             extra["name"],
                        "price":            extra_price,
                        "duration_minutes": extra_duration,
                    })

            # ── Convertir fecha/hora local a UTC (vía PostgreSQL) ──
            business_timezone    = business["timezone"] or "America/Monterrey"
            start_datetime_str   = f"{appointment_date} {start_time}:00"

            cur.execute(
                """
                SELECT
                    (%s::timestamp AT TIME ZONE %s) AS start_time_db,
                    ((%s::timestamp AT TIME ZONE %s)
                     + (%s || ' minutes')::interval) AS end_time_db
                """,
                (
                    start_datetime_str, business_timezone,
                    start_datetime_str, business_timezone,
                    total_duration,
                ),
            )
            time_row     = cur.fetchone()
            start_time_db = time_row["start_time_db"]
            end_time_db   = time_row["end_time_db"]

            overlap = find_overlapping_appointment(
                cur,
                business["id"],
                start_time_db,
                end_time_db,
                business_timezone,
                staff_id=staff_id,
            )
            if overlap:
                conn.rollback()
                flash(build_overlap_message(overlap, staff_id=staff_id), "warning")
                return redirect(url_for("nails.agenda"))

            # ── Insertar la cita ───────────────────────────
            cur.execute(
                """
                INSERT INTO nails_appointments (
                    business_id, client_id, staff_id, service_id,
                    title, start_time, end_time, status,
                    estimated_total, deposit_amount, notes
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    business["id"],
                    client_id if client_id else None,
                    staff_id  if staff_id  else None,
                    service_id,
                    " + ".join(item["name"] for item in services_to_insert[:3]),
                    start_time_db,
                    end_time_db,
                    status,
                    estimated_total,
                    deposit_amount,
                    notes,
                ),
            )
            appointment_id = cur.fetchone()["id"]

            # ── Insertar servicios de la cita ──────────────
            for service_item in services_to_insert:
                cur.execute(
                    """
                    INSERT INTO nails_appointment_services (
                        appointment_id, service_id, name, price, duration_minutes, sort_order
                    )
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (
                        appointment_id,
                        service_item["id"],
                        service_item["name"],
                        service_item["price"],
                        service_item["duration_minutes"],
                        service_item["sort_order"],
                    ),
                )

            # ── Insertar extras de la cita ─────────────────
            for extra in extras_to_insert:
                cur.execute(
                    """
                    INSERT INTO nails_appointment_extras (
                        appointment_id, extra_id, name, price, duration_minutes
                    )
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        appointment_id,
                        extra["id"],
                        extra["name"],
                        extra["price"],
                        extra["duration_minutes"],
                    ),
                )

            sale_detail_items = [
                {
                    "item_type": "service",
                    "item_id": item["id"],
                    "name": item["name"],
                    "description": None,
                    "quantity": 1,
                    "unit_price": item["price"],
                    "total": item["price"],
                }
                for item in services_to_insert
            ]
            sale_detail_items.extend(
                {
                    "item_type": "extra",
                    "item_id": extra["id"],
                    "name": extra["name"],
                    "description": None,
                    "quantity": 1,
                    "unit_price": extra["price"],
                    "total": extra["price"],
                }
                for extra in extras_to_insert
            )
            sync_sale_for_appointment(
                cur,
                business["id"],
                appointment_id,
                client_id,
                staff_id,
                sale_detail_items,
                estimated_total,
                initial_paid_amount=deposit_amount,
                notes=notes,
            )

            if status == "cancelada":
                cur.execute(
                    """
                    UPDATE nails_sales
                    SET status = 'cancelada',
                        balance_due = 0,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE appointment_id = %s
                      AND business_id = %s
                      AND status != 'cancelada'
                    """,
                    (appointment_id, business["id"]),
                )

            conn.commit()
            flash("Cita agregada correctamente con venta pendiente ligada.", "success")
            return redirect(url_for("nails.agenda"))

        # ── GET: cargar datos para renderizar la vista ─────
        cur.execute(
            """
            SELECT * FROM nails_clients
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY name ASC
            """,
            (business["id"],),
        )
        clients = cur.fetchall()

        cur.execute(
            """
            SELECT * FROM nails_staff
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY role ASC, name ASC
            """,
            (business["id"],),
        )
        staff = cur.fetchall()

        cur.execute(
            """
            SELECT s.*, c.name AS category_name
            FROM nails_services s
            LEFT JOIN nails_service_categories c ON c.id = s.category_id
            WHERE s.business_id = %s AND s.is_active = TRUE
            ORDER BY c.sort_order ASC, s.sort_order ASC, s.name ASC
            """,
            (business["id"],),
        )
        services = cur.fetchall()

        cur.execute(
            """
            SELECT * FROM nails_extras
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY sort_order ASC, name ASC
            """,
            (business["id"],),
        )
        extras = cur.fetchall()

        business_timezone = business["timezone"] or "America/Monterrey"

        # Citas recientes y futuras para calendario y validación de empalmes en modal.
        cur.execute(
            """
            SELECT
                a.*,
                c.name  AS client_name,
                c.phone AS client_phone,
                s.name  AS service_name,
                st.name AS staff_name,
                TO_CHAR(a.start_time AT TIME ZONE %s, 'YYYY-MM-DD"T"HH24:MI:SS') AS start_time_iso,
                TO_CHAR(a.end_time   AT TIME ZONE %s, 'YYYY-MM-DD"T"HH24:MI:SS') AS end_time_iso,
                TO_CHAR(a.start_time AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS start_time_formatted,
                TO_CHAR(a.end_time   AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS end_time_formatted
            FROM nails_appointments a
            LEFT JOIN nails_clients  c  ON c.id  = a.client_id
            LEFT JOIN nails_services s  ON s.id  = a.service_id
            LEFT JOIN nails_staff    st ON st.id = a.staff_id
            WHERE a.business_id = %s
              AND a.start_time >= NOW() - INTERVAL '90 days'
            ORDER BY a.start_time ASC
            LIMIT 500
            """,
            (
                business_timezone, business_timezone,
                business_timezone, business_timezone,
                business["id"],
            ),
        )
        appointments = cur.fetchall()

        cur.execute(
            """
            SELECT
                COUNT(*) FILTER (
                    WHERE DATE(start_time AT TIME ZONE %s) = DATE(NOW() AT TIME ZONE %s)
                      AND status NOT IN ('cancelada', 'no_asistio')
                ) AS appointments_today,
                COUNT(*) FILTER (
                    WHERE status = 'pendiente'
                      AND start_time >= NOW()
                ) AS pending_count,
                COALESCE(SUM(deposit_amount) FILTER (
                    WHERE COALESCE(deposit_amount, 0) > 0
                      AND status NOT IN ('cancelada', 'no_asistio')
                ), 0) AS deposits_total
            FROM nails_appointments
            WHERE business_id = %s
            """,
            (business_timezone, business_timezone, business["id"]),
        )
        agenda_stats_row = cur.fetchone()

        cur.execute(
            """
            SELECT COALESCE(SUM(total), 0) AS today_sales_total
            FROM nails_sales
            WHERE business_id = %s
              AND status != 'cancelada'
              AND DATE(created_at AT TIME ZONE %s) = DATE(NOW() AT TIME ZONE %s)
            """,
            (business["id"], business_timezone, business_timezone),
        )
        today_sales = cur.fetchone()

        cur.execute(
            """
            SELECT
                a.id,
                c.name AS client_name,
                TO_CHAR(a.start_time AT TIME ZONE %s, 'HH24:MI') AS start_time_label
            FROM nails_appointments a
            LEFT JOIN nails_clients c ON c.id = a.client_id
            WHERE a.business_id = %s
              AND a.start_time >= NOW()
              AND a.status IN ('pendiente', 'confirmada')
            ORDER BY a.start_time ASC
            LIMIT 1
            """,
            (business_timezone, business["id"]),
        )
        next_appointment = cur.fetchone()

        agenda_stats = {
            "appointments_today": agenda_stats_row["appointments_today"] or 0,
            "pending_count": agenda_stats_row["pending_count"] or 0,
            "deposits_total": agenda_stats_row["deposits_total"] or 0,
            "today_sales_total": today_sales["today_sales_total"] if today_sales else 0,
            "next_time": next_appointment["start_time_label"] if next_appointment else "Sin cita",
            "next_client": (next_appointment["client_name"] if next_appointment else "") or "Cliente General",
        }

        # Extras de todas las citas cargadas, agrupados por appointment_id
        cur.execute(
            """
            SELECT ae.appointment_id, ae.extra_id, ae.name, ae.price, ae.duration_minutes
            FROM nails_appointment_extras ae
            INNER JOIN nails_appointments a ON a.id = ae.appointment_id
            WHERE a.business_id = %s
            ORDER BY ae.id ASC
            """,
            (business["id"],),
        )
        appointment_extras_rows = cur.fetchall()

        appointment_extras: dict = {}
        for row in appointment_extras_rows:
            aid = row["appointment_id"]
            appointment_extras.setdefault(aid, []).append(row)

        cur.execute(
            """
            SELECT aps.appointment_id, aps.service_id, aps.name, aps.price, aps.duration_minutes
            FROM nails_appointment_services aps
            INNER JOIN nails_appointments a ON a.id = aps.appointment_id
            WHERE a.business_id = %s
            ORDER BY aps.sort_order ASC, aps.id ASC
            """,
            (business["id"],),
        )
        appointment_services_rows = cur.fetchall()

        appointment_services: dict = {}
        for row in appointment_services_rows:
            aid = row["appointment_id"]
            appointment_services.setdefault(aid, []).append(row)

        return render_template(
            "nails/agenda.html",
            business=business,
            clients=clients,
            staff=staff,
            services=services,
            extras=extras,
            appointments=appointments,
            agenda_stats=agenda_stats,
            appointment_extras=appointment_extras,
            appointment_services=appointment_services,
            can_delete_appointments=user_is_nails_owner(cur, user_id, business),
        )
        # BUG CORREGIDO: se eliminó el `return render_template(...)` duplicado
        # que había después del bloque finally y era código muerto inalcanzable.

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error en agenda: {e}", "danger")
        return redirect(url_for("nails.dashboard"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# CLIENTES
# GET:  Lista todas las clientas activas del salón.
# POST: Registra una nueva clienta con validación de email
#       y fecha de cumpleaños.
# =========================================================

@nails_bp.route("/clientes", methods=["GET", "POST"])
def clientes():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        business_timezone = business["timezone"] or "America/Monterrey"

        if request.method == "POST":
            name             = clean_text(request.form.get("name"), 120)
            phone            = clean_text(request.form.get("phone"), 40)
            email            = clean_text(request.form.get("email"), 120).lower()
            instagram        = clean_text(request.form.get("instagram"), 80)
            birthday         = clean_text(request.form.get("birthday"))
            preferences      = clean_text(request.form.get("preferences"), 1000)
            allergies_notes  = clean_text(request.form.get("allergies_notes"), 1000)
            notes            = clean_text(request.form.get("notes"), 1000)

            if not name:
                flash("El nombre de la clienta es obligatorio.", "warning")
                return redirect(url_for("nails.clientes"))

            if email and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
                flash("El correo de la clienta no es válido.", "warning")
                return redirect(url_for("nails.clientes"))

            birthday_value = parse_date_value(birthday)
            if birthday and not birthday_value:
                flash("La fecha de cumpleaños no es válida.", "warning")
                return redirect(url_for("nails.clientes"))

            cur.execute(
                """
                INSERT INTO nails_clients (
                    business_id, name, phone, email, instagram,
                    birthday, preferences, allergies_notes, notes
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    business["id"], name, phone, email, instagram,
                    birthday_value, preferences, allergies_notes, notes,
                ),
            )
            conn.commit()
            flash("Clienta agregada correctamente.", "success")
            return redirect(url_for("nails.clientes"))

        q = clean_text(request.args.get("q"), 80)
        client_filter = clean_text(request.args.get("filter"), 40)
        frequency_filter = clean_text(request.args.get("frequency"), 40)
        valid_client_filters = {"frecuentes", "cumpleanos", "primera_visita", "saldo_pendiente", "por_volver"}
        valid_frequency_filters = {"sin_visitas", "primera_visita", "frecuentes"}

        if client_filter not in valid_client_filters:
            client_filter = ""
        if frequency_filter not in valid_frequency_filters:
            frequency_filter = ""

        page = max(request.args.get("page", 1, type=int), 1)
        per_page = 25
        offset = (page - 1) * per_page
        like = f"%{q}%"
        where_sql = "WHERE c.business_id = %s AND c.is_active = TRUE"
        params = [business["id"]]

        if q:
            where_sql += """
                AND (
                    TRANSLATE(LOWER(c.name), 'áéíóú', 'aeiou') ILIKE TRANSLATE(LOWER(%s), 'áéíóú', 'aeiou')
                    OR COALESCE(c.phone, '') ILIKE %s
                    OR COALESCE(c.email, '') ILIKE %s
                    OR COALESCE(c.instagram, '') ILIKE %s
                )
            """
            params.extend([like, like, like, like])

        if client_filter == "frecuentes":
            where_sql += " AND COALESCE(a.appointments_count, c.total_visits, 0) >= 2"
        elif client_filter == "cumpleanos":
            where_sql += """
                AND c.birthday IS NOT NULL
                AND EXTRACT(MONTH FROM c.birthday) = EXTRACT(MONTH FROM NOW() AT TIME ZONE %s)
            """
            params.append(business_timezone)
        elif client_filter == "primera_visita":
            where_sql += " AND COALESCE(a.appointments_count, c.total_visits, 0) = 1"
        elif client_filter == "saldo_pendiente":
            where_sql += " AND COALESCE(v.balance_due_total, 0) > 0"
        elif client_filter == "por_volver":
            where_sql += """
                AND last_appt.last_visit IS NOT NULL
                AND last_appt.last_visit < NOW() - INTERVAL '30 days'
                AND next_appt.next_visit IS NULL
            """

        if frequency_filter == "sin_visitas":
            where_sql += " AND COALESCE(a.appointments_count, c.total_visits, 0) = 0"
        elif frequency_filter == "primera_visita":
            where_sql += " AND COALESCE(a.appointments_count, c.total_visits, 0) = 1"
        elif frequency_filter == "frecuentes":
            where_sql += " AND COALESCE(a.appointments_count, c.total_visits, 0) >= 2"

        cur.execute(
            f"""
            SELECT
                c.*,
                COUNT(*) OVER() AS total_filtered,
                COALESCE(v.sales_count, 0) AS sales_count,
                COALESCE(v.sales_total, c.total_spent, 0) AS total_spent_calculated,
                COALESCE(v.balance_due_total, 0) AS balance_due_total,
                COALESCE(a.appointments_count, c.total_visits, 0) AS visits_calculated,
                TO_CHAR(last_appt.last_visit AT TIME ZONE %s, 'DD Mon YYYY') AS last_visit_label,
                TO_CHAR(next_appt.next_visit AT TIME ZONE %s, 'DD Mon YYYY HH24:MI') AS next_visit_label,
                last_appt.last_visit,
                next_appt.next_visit,
                (
                    c.birthday IS NOT NULL
                    AND EXTRACT(MONTH FROM c.birthday) = EXTRACT(MONTH FROM NOW() AT TIME ZONE %s)
                ) AS is_birthday_month,
                (
                    last_appt.last_visit IS NOT NULL
                    AND last_appt.last_visit < NOW() - INTERVAL '30 days'
                    AND next_appt.next_visit IS NULL
                ) AS is_due_return
            FROM nails_clients c
            LEFT JOIN (
                SELECT
                    client_id,
                    COUNT(*) AS sales_count,
                    COALESCE(SUM(total), 0) AS sales_total,
                    COALESCE(SUM(balance_due), 0) AS balance_due_total
                FROM nails_sales
                WHERE business_id = %s AND status != 'cancelada' AND client_id IS NOT NULL
                GROUP BY client_id
            ) v ON v.client_id = c.id
            LEFT JOIN (
                SELECT client_id, COUNT(*) AS appointments_count
                FROM nails_appointments
                WHERE business_id = %s
                  AND status NOT IN ('cancelada', 'no_asistio')
                  AND start_time <= NOW()
                  AND client_id IS NOT NULL
                GROUP BY client_id
            ) a ON a.client_id = c.id
            LEFT JOIN (
                SELECT client_id, MAX(start_time) AS last_visit
                FROM nails_appointments
                WHERE business_id = %s
                  AND status NOT IN ('cancelada', 'no_asistio')
                  AND start_time <= NOW()
                  AND client_id IS NOT NULL
                GROUP BY client_id
            ) last_appt ON last_appt.client_id = c.id
            LEFT JOIN (
                SELECT client_id, MIN(start_time) AS next_visit
                FROM nails_appointments
                WHERE business_id = %s
                  AND status IN ('pendiente', 'confirmada')
                  AND start_time >= NOW()
                  AND client_id IS NOT NULL
                GROUP BY client_id
            ) next_appt ON next_appt.client_id = c.id
            {where_sql}
            ORDER BY LOWER(c.name) ASC
            LIMIT %s OFFSET %s
            """,
            [
                business_timezone,
                business_timezone,
                business_timezone,
                business["id"],
                business["id"],
                business["id"],
                business["id"],
                *params,
                per_page,
                offset,
            ],
        )
        clients = cur.fetchall()
        total_filtered = clients[0]["total_filtered"] if clients else 0
        total_pages = max(1, (total_filtered + per_page - 1) // per_page)

        cur.execute(
            """
            WITH visit_stats AS (
                SELECT
                    client_id,
                    COUNT(*) AS visits_count,
                    MAX(start_time) AS last_visit
                FROM nails_appointments
                WHERE business_id = %s
                  AND status NOT IN ('cancelada', 'no_asistio')
                  AND start_time <= NOW()
                  AND client_id IS NOT NULL
                GROUP BY client_id
            ),
            next_stats AS (
                SELECT client_id, MIN(start_time) AS next_visit
                FROM nails_appointments
                WHERE business_id = %s
                  AND status IN ('pendiente', 'confirmada')
                  AND start_time >= NOW()
                  AND client_id IS NOT NULL
                GROUP BY client_id
            )
            SELECT
                COUNT(*) AS total_clients,
                COUNT(*) FILTER (WHERE COALESCE(v.visits_count, c.total_visits, 0) >= 2) AS frequent_clients,
                COUNT(*) FILTER (
                    WHERE c.birthday IS NOT NULL
                      AND EXTRACT(MONTH FROM c.birthday) = EXTRACT(MONTH FROM NOW() AT TIME ZONE %s)
                ) AS birthdays_month,
                COUNT(*) FILTER (
                    WHERE v.last_visit IS NOT NULL
                      AND v.last_visit < NOW() - INTERVAL '30 days'
                      AND n.next_visit IS NULL
                ) AS due_return
            FROM nails_clients c
            LEFT JOIN visit_stats v ON v.client_id = c.id
            LEFT JOIN next_stats n ON n.client_id = c.id
            WHERE c.business_id = %s AND c.is_active = TRUE
            """,
            (business["id"], business["id"], business_timezone, business["id"]),
        )
        stats_row = cur.fetchone()

        stats = {
            "total_clients": stats_row["total_clients"] or 0,
            "frequent_clients": stats_row["frequent_clients"] or 0,
            "birthdays_month": stats_row["birthdays_month"] or 0,
            "due_return": stats_row["due_return"] or 0,
        }

        return render_template(
            "nails/clientes.html",
            business=business,
            clients=clients,
            stats=stats,
            q=q,
            client_filter=client_filter,
            frequency_filter=frequency_filter,
            page=page,
            per_page=per_page,
            total_filtered=total_filtered,
            total_pages=total_pages,
        )

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error en clientas: {e}", "danger")
        return redirect(url_for("nails.dashboard"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# SERVICIOS Y EXTRAS
# GET:  Lista todos los servicios y extras activos del salón.
# POST: Crea un nuevo servicio (form_type='service') o un
#       extra (form_type='extra'). Si la categoría del servicio
#       no existe se crea automáticamente.
# =========================================================

@nails_bp.route("/servicios", methods=["GET", "POST"])
def servicios():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        ensure_nails_service_icon_column(cur)
        conn.commit()

        if request.method == "POST":
            form_type = request.form.get("form_type")

            # ── Alta de servicio ───────────────────────────
            if form_type == "service":
                name               = clean_text(request.form.get("name"), 160)
                description        = clean_text(request.form.get("description"), 1200)
                category_name      = clean_text(request.form.get("category"), 80)
                base_price_raw     = request.form.get("base_price", "0")
                duration_minutes_raw = request.form.get("duration_minutes", "60")
                is_public          = request.form.get("is_public") == "on"
                service_icon       = clean_service_icon(request.form.get("service_icon"), category_name)

                if not name:
                    flash("El nombre del servicio es obligatorio.", "warning")
                    return redirect(url_for("nails.servicios"))

                base_price       = parse_positive_float(base_price_raw)
                duration_minutes = parse_positive_int(duration_minutes_raw, default=60, max_value=1440)

                # Buscar o crear categoría por nombre (case-insensitive)
                category_id = None
                if category_name:
                    cur.execute(
                        """
                        SELECT id FROM nails_service_categories
                        WHERE business_id = %s AND LOWER(name) = LOWER(%s)
                        LIMIT 1
                        """,
                        (business["id"], category_name),
                    )
                    existing_category = cur.fetchone()

                    if existing_category:
                        category_id = existing_category["id"]
                    else:
                        cur.execute(
                            """
                            INSERT INTO nails_service_categories (business_id, name)
                            VALUES (%s, %s)
                            RETURNING id
                            """,
                            (business["id"], category_name),
                        )
                        category_id = cur.fetchone()["id"]

                cur.execute(
                    """
                    INSERT INTO nails_services (
                        business_id, category_id, name, description,
                        base_price, duration_minutes, is_public, service_icon
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        business["id"], category_id, name, description,
                        base_price, duration_minutes, is_public, service_icon,
                    ),
                )
                conn.commit()
                flash("Servicio agregado correctamente.", "success")
                return redirect(url_for("nails.servicios"))

            # ── Alta de extra ──────────────────────────────
            if form_type == "extra":
                name                 = clean_text(request.form.get("extra_name"), 160)
                description          = clean_text(request.form.get("extra_description"), 1200)
                price_raw            = request.form.get("extra_price", "0")
                duration_minutes_raw = request.form.get("extra_duration_minutes", "0")

                if not name:
                    flash("El nombre del extra es obligatorio.", "warning")
                    return redirect(url_for("nails.servicios"))

                price            = parse_positive_float(price_raw)
                duration_minutes = parse_positive_int(duration_minutes_raw, default=0, max_value=1440)

                cur.execute(
                    """
                    INSERT INTO nails_extras (
                        business_id, name, description, price, duration_minutes
                    )
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (business["id"], name, description, price, duration_minutes),
                )
                conn.commit()
                flash("Extra agregado correctamente.", "success")
                return redirect(url_for("nails.servicios"))

        # ── GET: listar servicios y extras paginados ───────
        service_page = max(request.args.get("service_page", 1, type=int), 1)
        extra_page = max(request.args.get("extra_page", 1, type=int), 1)
        services_per_page = 6
        extras_per_page = 4
        service_offset = (service_page - 1) * services_per_page
        extra_offset = (extra_page - 1) * extras_per_page

        cur.execute(
            """
            SELECT COUNT(*) AS total
            FROM nails_services
            WHERE business_id = %s AND is_active = TRUE
            """,
            (business["id"],),
        )
        services_total = cur.fetchone()["total"] or 0
        services_total_pages = max(1, (services_total + services_per_page - 1) // services_per_page)

        cur.execute(
            """
            SELECT COUNT(*) AS total
            FROM nails_extras
            WHERE business_id = %s AND is_active = TRUE
            """,
            (business["id"],),
        )
        extras_total = cur.fetchone()["total"] or 0
        extras_total_pages = max(1, (extras_total + extras_per_page - 1) // extras_per_page)

        cur.execute(
            """
            SELECT s.*, c.name AS category_name
            FROM nails_services s
            LEFT JOIN nails_service_categories c ON c.id = s.category_id
            WHERE s.business_id = %s AND s.is_active = TRUE
            ORDER BY c.sort_order ASC, s.sort_order ASC, s.name ASC
            LIMIT %s OFFSET %s
            """,
            (business["id"], services_per_page, service_offset),
        )
        services = cur.fetchall()

        cur.execute(
            """
            SELECT * FROM nails_extras
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY sort_order ASC, name ASC
            LIMIT %s OFFSET %s
            """,
            (business["id"], extras_per_page, extra_offset),
        )
        extras = cur.fetchall()

        cur.execute(
            """
            SELECT name
            FROM nails_service_categories
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY sort_order ASC, name ASC
            """,
            (business["id"],),
        )
        categories = cur.fetchall()

        return render_template(
            "nails/servicios.html",
            business=business,
            services=services,
            extras=extras,
            categories=categories,
            service_icon_options=NAILS_SERVICE_ICON_OPTIONS,
            service_page=service_page,
            extra_page=extra_page,
            services_total=services_total,
            extras_total=extras_total,
            services_total_pages=services_total_pages,
            extras_total_pages=extras_total_pages,
            services_per_page=services_per_page,
            extras_per_page=extras_per_page,
        )

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error en servicios: {e}", "danger")
        return redirect(url_for("nails.dashboard"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/servicios/<int:service_id>/editar", methods=["POST"])
def editar_servicio(service_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        ensure_nails_service_icon_column(cur)
        conn.commit()

        cur.execute(
            """
            SELECT id
            FROM nails_services
            WHERE id = %s AND business_id = %s AND is_active = TRUE
            LIMIT 1
            """,
            (service_id, business["id"]),
        )
        if not cur.fetchone():
            flash("El servicio no existe o ya fue eliminado.", "warning")
            return redirect(url_for("nails.servicios"))

        name                 = clean_text(request.form.get("name"), 160)
        description          = clean_text(request.form.get("description"), 1200)
        category_name        = clean_text(request.form.get("category"), 80)
        base_price           = parse_positive_float(request.form.get("base_price", "0"))
        duration_minutes     = parse_positive_int(request.form.get("duration_minutes", "60"), default=60, max_value=1440)
        is_public            = request.form.get("is_public") == "on"
        service_icon         = clean_service_icon(request.form.get("service_icon"), category_name)

        if not name:
            flash("El nombre del servicio es obligatorio.", "warning")
            return redirect(url_for("nails.servicios"))

        category_id = None
        if category_name:
            cur.execute(
                """
                SELECT id
                FROM nails_service_categories
                WHERE business_id = %s AND LOWER(name) = LOWER(%s)
                LIMIT 1
                """,
                (business["id"], category_name),
            )
            existing_category = cur.fetchone()

            if existing_category:
                category_id = existing_category["id"]
            else:
                cur.execute(
                    """
                    INSERT INTO nails_service_categories (business_id, name)
                    VALUES (%s, %s)
                    RETURNING id
                    """,
                    (business["id"], category_name),
                )
                category_id = cur.fetchone()["id"]

        cur.execute(
            """
            UPDATE nails_services
            SET category_id = %s,
                name = %s,
                description = %s,
                base_price = %s,
                duration_minutes = %s,
                is_public = %s,
                service_icon = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (
                category_id, name, description, base_price, duration_minutes,
                is_public, service_icon, service_id, business["id"],
            ),
        )

        conn.commit()
        flash("Servicio actualizado correctamente.", "success")
        return redirect(url_for("nails.servicios"))

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error al editar el servicio: {e}", "danger")
        return redirect(url_for("nails.servicios"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/servicios/<int:service_id>/eliminar", methods=["POST"])
def eliminar_servicio(service_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            UPDATE nails_services
            SET is_active = FALSE,
                is_public = FALSE,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s AND is_active = TRUE
            """,
            (service_id, business["id"]),
        )
        conn.commit()

        if cur.rowcount:
            flash("Servicio eliminado correctamente.", "success")
        else:
            flash("El servicio no existe o ya fue eliminado.", "warning")

        return redirect(url_for("nails.servicios"))

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error al eliminar el servicio: {e}", "danger")
        return redirect(url_for("nails.servicios"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/extras/<int:extra_id>/editar", methods=["POST"])
def editar_extra(extra_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            SELECT id
            FROM nails_extras
            WHERE id = %s AND business_id = %s AND is_active = TRUE
            LIMIT 1
            """,
            (extra_id, business["id"]),
        )
        if not cur.fetchone():
            flash("El extra no existe o ya fue eliminado.", "warning")
            return redirect(url_for("nails.servicios"))

        name             = clean_text(request.form.get("extra_name"), 160)
        description      = clean_text(request.form.get("extra_description"), 1200)
        price            = parse_positive_float(request.form.get("extra_price", "0"))
        duration_minutes = parse_positive_int(request.form.get("extra_duration_minutes", "0"), default=0, max_value=1440)

        if not name:
            flash("El nombre del extra es obligatorio.", "warning")
            return redirect(url_for("nails.servicios"))

        cur.execute(
            """
            UPDATE nails_extras
            SET name = %s,
                description = %s,
                price = %s,
                duration_minutes = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (name, description, price, duration_minutes, extra_id, business["id"]),
        )

        conn.commit()
        flash("Extra actualizado correctamente.", "success")
        return redirect(url_for("nails.servicios"))

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error al editar el extra: {e}", "danger")
        return redirect(url_for("nails.servicios"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/extras/<int:extra_id>/eliminar", methods=["POST"])
def eliminar_extra(extra_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            UPDATE nails_extras
            SET is_active = FALSE,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s AND is_active = TRUE
            """,
            (extra_id, business["id"]),
        )
        conn.commit()

        if cur.rowcount:
            flash("Extra eliminado correctamente.", "success")
        else:
            flash("El extra no existe o ya fue eliminado.", "warning")

        return redirect(url_for("nails.servicios"))

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error al eliminar el extra: {e}", "danger")
        return redirect(url_for("nails.servicios"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# VENTAS
# GET:  Muestra formulario de nueva venta y las últimas 30.
#       Si se pasa ?appointment_id o ?nueva_desde_cita precarga
#       la cita en el formulario.
# POST: Registra la venta, sus detalles, el pago inicial,
#       actualiza la cita a 'atendida' y los contadores de la clienta.
# =========================================================

@nails_bp.route("/ventas", methods=["GET", "POST"])
def ventas():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        ensure_nails_appointment_services_table(cur)
        conn.commit()

        if request.method == "POST":
            # ── Leer IDs del formulario ────────────────────
            client_id      = clean_optional_id(request.form.get("client_id"))
            appointment_id = clean_optional_id(request.form.get("appointment_id"))
            staff_id       = clean_optional_id(request.form.get("staff_id"))
            selected_service_ids = [
                clean_optional_id(value)
                for value in request.form.getlist("service_ids")
            ]
            selected_service_ids = [value for value in selected_service_ids if value]
            service_id = clean_optional_id(request.form.get("service_id"))

            if not selected_service_ids and service_id:
                selected_service_ids = [service_id]
            service_id = selected_service_ids[0] if selected_service_ids else None

            # Si se vincula una cita y no se eligió clienta manualmente,
            # heredar clienta y técnica de la cita.
            if appointment_id and not client_id:
                cur.execute(
                    """
                    SELECT client_id, staff_id
                    FROM nails_appointments
                    WHERE id = %s AND business_id = %s
                    LIMIT 1
                    """,
                    (appointment_id, business["id"]),
                )
                appointment_data = cur.fetchone()
                if appointment_data:
                    client_id = appointment_data["client_id"] or client_id
                    staff_id  = appointment_data["staff_id"]  or staff_id

            selected_extras      = request.form.getlist("extras")
            discount_amount_raw  = request.form.get("discount_amount", "0")
            paid_amount_raw      = request.form.get("paid_amount", "0")
            payment_method       = clean_text(request.form.get("payment_method")) or None
            notes                = clean_text(request.form.get("notes"), 1000)
            after_save_action    = clean_text(request.form.get("after_save_action"), 30)

            # ── Validaciones básicas ───────────────────────
            if not service_id:
                flash("Selecciona un servicio para registrar la venta.", "warning")
                return redirect(url_for("nails.ventas"))

            if payment_method and payment_method not in PAYMENT_METHODS:
                flash("Método de pago inválido.", "warning")
                return redirect(url_for("nails.ventas"))

            # ── Validar ownership de IDs ───────────────────
            if appointment_id and not row_belongs_to_business(
                cur, "nails_appointments", appointment_id, business["id"], active_only=False
            ):
                flash("La cita seleccionada no pertenece a este salón.", "warning")
                return redirect(url_for("nails.ventas"))

            if client_id and not row_belongs_to_business(cur, "nails_clients", client_id, business["id"]):
                flash("La clienta seleccionada no pertenece a este salón.", "warning")
                return redirect(url_for("nails.ventas"))

            if staff_id and not row_belongs_to_business(cur, "nails_staff", staff_id, business["id"]):
                flash("La técnica seleccionada no pertenece a este salón.", "warning")
                return redirect(url_for("nails.ventas"))

            # ── Servicios de la venta ──────────────────────
            cur.execute(
                """
                SELECT id, name, base_price
                FROM nails_services
                WHERE id = ANY(%s)
                  AND business_id = %s
                  AND is_active = TRUE
                ORDER BY array_position(%s, id)
                """,
                (selected_service_ids, business["id"], selected_service_ids),
            )
            service_rows = cur.fetchall()

            if len(service_rows) != len(set(selected_service_ids)):
                flash("Uno o más servicios seleccionados no existen.", "warning")
                return redirect(url_for("nails.ventas"))

            subtotal = 0
            detail_items = []

            for service in service_rows:
                service_price = float(service["base_price"] or 0)
                subtotal += service_price
                detail_items.append({
                    "item_type":   "service",
                    "item_id":     service["id"],
                    "name":        service["name"],
                    "description": None,
                    "quantity":    1,
                    "unit_price":  service_price,
                    "total":       service_price,
                })

            # ── Extras de la venta ─────────────────────────
            for extra_id in selected_extras:
                cur.execute(
                    """
                    SELECT id, name, description, price
                    FROM nails_extras
                    WHERE id = %s AND business_id = %s AND is_active = TRUE
                    LIMIT 1
                    """,
                    (extra_id, business["id"]),
                )
                extra = cur.fetchone()

                if extra:
                    extra_price = float(extra["price"] or 0)
                    subtotal   += extra_price
                    detail_items.append({
                        "item_type":   "extra",
                        "item_id":     extra["id"],
                        "name":        extra["name"],
                        "description": extra["description"],
                        "quantity":    1,
                        "unit_price":  extra_price,
                        "total":       extra_price,
                    })

            # ── Calcular totales ───────────────────────────
            discount_amount = parse_positive_float(discount_amount_raw, max_value=subtotal)
            tax_amount      = 0
            total           = subtotal - discount_amount + tax_amount
            paid_amount     = parse_positive_float(paid_amount_raw, max_value=total)
            balance_due     = total - paid_amount

            if subtotal <= 0 or total <= 0 or not detail_items:
                flash("Selecciona un servicio con importe para poder cobrar.", "warning")
                return redirect(url_for("nails.ventas"))

            # Determinar estado según montos
            if total <= 0:
                status = "pagada"
            elif paid_amount <= 0:
                status = "pendiente"
            elif balance_due > 0:
                status = "anticipo"
            else:
                status = "pagada"

            existing_sale = None
            if appointment_id:
                cur.execute(
                    """
                    SELECT *
                    FROM nails_sales
                    WHERE business_id = %s AND appointment_id = %s
                    ORDER BY id ASC
                    LIMIT 1
                    """,
                    (business["id"], appointment_id),
                )
                existing_sale = cur.fetchone()

            previous_paid_amount = 0

            if existing_sale:
                sale_id = existing_sale["id"]
                previous_paid_amount = float(existing_sale["paid_amount"] or 0)

                cur.execute(
                    """
                    UPDATE nails_sales
                    SET client_id = %s,
                        staff_id = %s,
                        subtotal = %s,
                        discount_amount = %s,
                        discount_percentage = 0,
                        tax_amount = %s,
                        total = %s,
                        paid_amount = %s,
                        balance_due = %s,
                        payment_method = %s,
                        status = %s,
                        notes = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s AND business_id = %s
                    """,
                    (
                        client_id, staff_id, subtotal, discount_amount,
                        tax_amount, total, paid_amount, balance_due,
                        payment_method, status, notes, sale_id, business["id"],
                    ),
                )
                cur.execute("DELETE FROM nails_sale_details WHERE sale_id = %s", (sale_id,))
            else:
                # ── Insertar venta ─────────────────────────────
                cur.execute(
                    """
                    INSERT INTO nails_sales (
                        business_id, client_id, appointment_id, staff_id,
                        subtotal, discount_amount, discount_percentage, tax_amount,
                        total, paid_amount, balance_due,
                        payment_method, status, notes
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        business["id"], client_id, appointment_id, staff_id,
                        subtotal, discount_amount, 0, tax_amount,
                        total, paid_amount, balance_due,
                        payment_method, status, notes,
                    ),
                )
                sale_id     = cur.fetchone()["id"]
                sale_number = f"N-{sale_id:06d}"

                cur.execute(
                    "UPDATE nails_sales SET sale_number = %s WHERE id = %s",
                    (sale_number, sale_id),
                )

            # ── Insertar detalles de la venta ──────────────
            for item in detail_items:
                cur.execute(
                    """
                    INSERT INTO nails_sale_details (
                        sale_id, item_type, item_id, name,
                        description, quantity, unit_price, total
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        sale_id,
                        item["item_type"],
                        item["item_id"],
                        item["name"],
                        item["description"],
                        item["quantity"],
                        item["unit_price"],
                        item["total"],
                    ),
                )

            # ── Registrar solo el pago incremental si lo hay ─
            paid_difference = round(paid_amount - previous_paid_amount, 2)
            if paid_difference > 0:
                payment_type = "anticipo" if status == "anticipo" else "pago"
                cur.execute(
                    """
                    INSERT INTO nails_payments (
                        sale_id, amount, payment_method, payment_type, notes
                    )
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        sale_id,
                        paid_difference,
                        payment_method or "efectivo",
                        payment_type,
                        "Pago registrado al actualizar venta de cita" if existing_sale else "Pago registrado al crear venta",
                    ),
                )

            # ── Actualizar contadores de la clienta ────────
            should_update_client_totals = (
                client_id
                and (
                    not existing_sale
                    or (existing_sale["status"] != "pagada" and status == "pagada")
                )
            )
            if should_update_client_totals:
                cur.execute(
                    """
                    UPDATE nails_clients
                    SET total_visits = COALESCE(total_visits, 0) + 1,
                        total_spent  = COALESCE(total_spent,  0) + %s,
                        updated_at   = CURRENT_TIMESTAMP
                    WHERE id = %s AND business_id = %s
                    """,
                    (total, client_id, business["id"]),
                )

            conn.commit()
            flash("Venta registrada correctamente.", "success")
            if after_save_action == "ticket":
                return redirect(url_for("nails.ticket", sale_id=sale_id))
            return redirect(url_for("nails.ventas", sale_id=sale_id))

        # ── GET: precargar cita si viene desde agenda ──────
        appointment_id = clean_optional_id(
            request.args.get("appointment_id") or request.args.get("nueva_desde_cita")
        )
        selected_appointment        = None
        selected_appointment_extras = []
        selected_appointment_services = []
        selected_appointment_sale   = None
        selected_extra_ids          = []
        selected_service_ids        = []

        if appointment_id:
            cur.execute(
                """
                SELECT
                    a.*,
                    c.name       AS client_name,
                    c.phone      AS client_phone,
                    s.name       AS service_name,
                    s.base_price AS service_price,
                    st.name      AS staff_name
                FROM nails_appointments a
                LEFT JOIN nails_clients  c  ON c.id  = a.client_id
                LEFT JOIN nails_services s  ON s.id  = a.service_id
                LEFT JOIN nails_staff    st ON st.id = a.staff_id
                WHERE a.id = %s AND a.business_id = %s
                LIMIT 1
                """,
                (appointment_id, business["id"]),
            )
            selected_appointment = cur.fetchone()

            if selected_appointment:
                cur.execute(
                    """
                    SELECT service_id, name, price, duration_minutes
                    FROM nails_appointment_services
                    WHERE appointment_id = %s
                    ORDER BY sort_order ASC, id ASC
                    """,
                    (appointment_id,),
                )
                selected_appointment_services = cur.fetchall()
                selected_service_ids = [
                    row["service_id"]
                    for row in selected_appointment_services
                    if row["service_id"]
                ]

                if not selected_service_ids and selected_appointment["service_id"]:
                    selected_service_ids = [selected_appointment["service_id"]]

                cur.execute(
                    """
                    SELECT extra_id, name, price
                    FROM nails_appointment_extras
                    WHERE appointment_id = %s
                    """,
                    (appointment_id,),
                )
                selected_appointment_extras = cur.fetchall()
                selected_extra_ids = [
                    row["extra_id"]
                    for row in selected_appointment_extras
                    if row["extra_id"]
                ]

                cur.execute(
                    """
                    SELECT *
                    FROM nails_sales
                    WHERE business_id = %s AND appointment_id = %s
                    ORDER BY id ASC
                    LIMIT 1
                    """,
                    (business["id"], appointment_id),
                )
                selected_appointment_sale = cur.fetchone()

        cur.execute(
            """
            SELECT * FROM nails_clients
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY name ASC
            """,
            (business["id"],),
        )
        clients = cur.fetchall()

        cur.execute(
            """
            SELECT * FROM nails_staff
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY role ASC, name ASC
            """,
            (business["id"],),
        )
        staff = cur.fetchall()

        cur.execute(
            """
            SELECT * FROM nails_services
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY name ASC
            """,
            (business["id"],),
        )
        services = cur.fetchall()

        cur.execute(
            """
            SELECT * FROM nails_extras
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY name ASC
            """,
            (business["id"],),
        )
        extras = cur.fetchall()

        business_timezone = business["timezone"] or "America/Monterrey"
        sales_q = clean_text(request.args.get("sales_q"), 120)
        sales_date = parse_date_value(request.args.get("sales_date"))
        sales_page = max(1, request.args.get("sales_page", 1, type=int) or 1)
        sales_per_page = 10
        sales_offset = (sales_page - 1) * sales_per_page

        sales_where = ["v.business_id = %s"]
        sales_params = [business["id"]]

        if sales_q:
            sales_where.append(
                """
                (
                    LOWER(COALESCE(v.sale_number, '')) LIKE LOWER(%s)
                    OR LOWER(COALESCE(c.name, 'Cliente General')) LIKE LOWER(%s)
                    OR LOWER(COALESCE(c.phone, '')) LIKE LOWER(%s)
                    OR LOWER(COALESCE(st.name, '')) LIKE LOWER(%s)
                )
                """
            )
            like_q = f"%{sales_q}%"
            sales_params.extend([like_q, like_q, like_q, like_q])

        if sales_date:
            sales_where.append("DATE(v.created_at AT TIME ZONE %s) = %s")
            sales_params.extend([business_timezone, sales_date])

        sales_where_sql = " AND ".join(sales_where)

        cur.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM nails_sales v
            LEFT JOIN nails_clients c  ON c.id  = v.client_id
            LEFT JOIN nails_staff   st ON st.id = v.staff_id
            WHERE {sales_where_sql}
            """,
            sales_params,
        )
        sales_total = cur.fetchone()["total"] or 0
        sales_total_pages = max(1, (sales_total + sales_per_page - 1) // sales_per_page)
        if sales_page > sales_total_pages:
            sales_page = sales_total_pages
            sales_offset = (sales_page - 1) * sales_per_page

        cur.execute(
            f"""
            SELECT
                v.*,
                c.name  AS client_name,
                c.phone AS client_phone,
                st.name AS staff_name,
                TO_CHAR(v.created_at AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS created_at_formatted
            FROM nails_sales v
            LEFT JOIN nails_clients c  ON c.id  = v.client_id
            LEFT JOIN nails_staff   st ON st.id = v.staff_id
            WHERE {sales_where_sql}
            ORDER BY v.created_at DESC, v.id DESC
            LIMIT %s OFFSET %s
            """,
            [business_timezone] + sales_params + [sales_per_page, sales_offset],
        )
        sales = cur.fetchall()

        highlight_sale_id = request.args.get("sale_id")
        can_cancel_sales = user_is_nails_owner(cur, user_id, business)

        return render_template(
            "nails/ventas.html",
            business=business,
            clients=clients,
            staff=staff,
            services=services,
            extras=extras,
            sales=sales,
            selected_appointment=selected_appointment,
            selected_appointment_sale=selected_appointment_sale,
            selected_appointment_services=selected_appointment_services,
            # BUG CORREGIDO: `selected_appointment_extras` antes no se pasaba
            # al template, por lo que la vista no podía mostrar los extras
            # de la cita preseleccionada.
            selected_appointment_extras=selected_appointment_extras,
            selected_service_ids=selected_service_ids,
            selected_extra_ids=selected_extra_ids,
            highlight_sale_id=highlight_sale_id,
            can_cancel_sales=can_cancel_sales,
            sales_q=sales_q,
            sales_date=sales_date.isoformat() if sales_date else "",
            sales_page=sales_page,
            sales_per_page=sales_per_page,
            sales_total=sales_total,
            sales_total_pages=sales_total_pages,
        )

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error en ventas: {e}", "danger")
        return redirect(url_for("nails.dashboard"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/ventas/<int:sale_id>/editar", methods=["POST"])
def editar_venta(sale_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            SELECT *
            FROM nails_sales
            WHERE id = %s AND business_id = %s
            LIMIT 1
            """,
            (sale_id, business["id"]),
        )
        sale = cur.fetchone()

        if not sale:
            flash("La venta no existe o no pertenece a este salón.", "warning")
            return redirect(url_for("nails.ventas"))

        if sale["status"] in ("pagada", "cancelada"):
            flash("Las ventas pagadas o canceladas solo se pueden consultar.", "info")
            return redirect(url_for("nails.ventas", sale_id=sale_id))

        client_id       = clean_optional_id(request.form.get("client_id"))
        staff_id        = clean_optional_id(request.form.get("staff_id"))
        payment_method  = clean_text(request.form.get("payment_method")) or None
        status          = clean_text(request.form.get("status"), 30) or sale["status"]
        notes           = clean_text(request.form.get("notes"), 1000)

        if payment_method and payment_method not in PAYMENT_METHODS:
            flash("Método de pago inválido.", "warning")
            return redirect(url_for("nails.ventas", sale_id=sale_id))

        if status not in SALE_STATUSES:
            flash("Estado de venta inválido.", "warning")
            return redirect(url_for("nails.ventas", sale_id=sale_id))

        if client_id and not row_belongs_to_business(cur, "nails_clients", client_id, business["id"]):
            flash("La clienta seleccionada no pertenece a este salón.", "warning")
            return redirect(url_for("nails.ventas", sale_id=sale_id))

        if staff_id and not row_belongs_to_business(cur, "nails_staff", staff_id, business["id"]):
            flash("La técnica seleccionada no pertenece a este salón.", "warning")
            return redirect(url_for("nails.ventas", sale_id=sale_id))

        subtotal        = float(sale["subtotal"] or 0)
        tax_amount      = float(sale["tax_amount"] or 0)
        discount_amount = parse_positive_float(
            request.form.get("discount_amount", sale["discount_amount"] or 0),
            max_value=subtotal,
        )
        total           = max(subtotal - discount_amount + tax_amount, 0)
        paid_amount     = parse_positive_float(
            request.form.get("paid_amount", sale["paid_amount"] or 0),
            max_value=total,
        )
        balance_due     = max(total - paid_amount, 0)

        if status != "cancelada":
            if total <= 0 or paid_amount >= total:
                status = "pagada"
            elif paid_amount > 0:
                status = "anticipo"
            else:
                status = "pendiente"

        cur.execute(
            """
            UPDATE nails_sales
            SET client_id = %s,
                staff_id = %s,
                discount_amount = %s,
                total = %s,
                paid_amount = %s,
                balance_due = %s,
                payment_method = %s,
                status = %s,
                notes = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (
                client_id, staff_id, discount_amount, total, paid_amount,
                balance_due, payment_method, status, notes, sale_id, business["id"],
            ),
        )

        previous_paid_amount = float(sale["paid_amount"] or 0)
        paid_difference = round(paid_amount - previous_paid_amount, 2)

        if paid_difference > 0 and status != "cancelada":
            cur.execute(
                """
                INSERT INTO nails_payments (sale_id, amount, payment_method, payment_type, notes)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    sale_id,
                    paid_difference,
                    payment_method or "efectivo",
                    "anticipo" if balance_due > 0 else "pago",
                    "Pago ajustado al editar venta",
                ),
            )
        elif paid_difference < 0:
            cur.execute(
                """
                INSERT INTO nails_payments (sale_id, amount, payment_method, payment_type, notes)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    sale_id,
                    abs(paid_difference),
                    payment_method or "efectivo",
                    "reembolso",
                    "Reembolso registrado al editar venta",
                ),
            )

        conn.commit()
        flash("Venta actualizada correctamente.", "success")
        return redirect(url_for("nails.ventas", sale_id=sale_id))

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error al editar la venta: {e}", "danger")
        return redirect(url_for("nails.ventas", sale_id=sale_id))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/ventas/<int:sale_id>/anular", methods=["POST"])
def anular_venta(sale_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        if not user_is_nails_owner(cur, user_id, business):
            flash("Solo la Jefa puede anular ventas.", "warning")
            return redirect(url_for("nails.ventas"))

        cur.execute(
            """
            SELECT *
            FROM nails_sales
            WHERE id = %s AND business_id = %s
            LIMIT 1
            """,
            (sale_id, business["id"]),
        )
        sale = cur.fetchone()

        if not sale:
            flash("La venta no existe o no pertenece a este salón.", "warning")
            return redirect(url_for("nails.ventas"))

        if sale["status"] == "cancelada":
            flash("Esta venta ya estaba anulada.", "info")
            return redirect(url_for("nails.ventas", sale_id=sale_id))

        original_status = sale["status"]
        original_total = float(sale["total"] or 0)
        client_id = sale["client_id"]
        current_notes = clean_text(sale["notes"], 1000)
        cancel_note = f"Venta anulada por Jefa el {datetime.now().strftime('%Y-%m-%d %H:%M')}."
        notes = (current_notes + "\n" + cancel_note).strip() if current_notes else cancel_note

        cur.execute(
            """
            UPDATE nails_sales
            SET status = 'cancelada',
                balance_due = 0,
                notes = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (notes, sale_id, business["id"]),
        )

        if client_id and original_status == "pagada":
            cur.execute(
                """
                UPDATE nails_clients
                SET total_visits = GREATEST(COALESCE(total_visits, 0) - 1, 0),
                    total_spent = GREATEST(COALESCE(total_spent, 0) - %s, 0),
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND business_id = %s
                """,
                (original_total, client_id, business["id"]),
            )

        cur.execute(
            """
            INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (
                business["id"],
                user_id,
                "sale_cancel",
                "Ventas",
                f"Anuló venta #{sale_id} ({sale['sale_number'] or 'sin folio'})",
            ),
        )

        conn.commit()
        flash("Venta anulada correctamente. Se conserva el historial y queda fuera de reportes activos.", "success")

    except Exception as e:
        conn.rollback()
        current_app.logger.error(f"NAILS_CANCEL_SALE_ERROR: {e}")
        flash(f"No se pudo anular la venta: {e}", "danger")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("nails.ventas", sale_id=sale_id))


# =========================================================
# GASTOS
# GET: Lista gastos del salón con filtros, KPIs y recurrentes.
# POST: Registra un gasto operativo para reportes más exactos.
# =========================================================

@nails_bp.route("/gastos", methods=["GET", "POST"])
def gastos():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur = conn.cursor()

    category_labels = {
        "materiales": "Materiales",
        "renta": "Renta",
        "servicios": "Servicios",
        "sueldos": "Sueldos",
        "comisiones": "Comisiones",
        "publicidad": "Publicidad",
        "mantenimiento": "Mantenimiento",
        "capacitacion": "Capacitación",
        "otros": "Otros",
    }
    payment_labels = {
        "efectivo": "Efectivo",
        "transferencia": "Transferencia",
        "tarjeta": "Tarjeta",
        "mixto": "Mixto",
        "otro": "Otro",
    }
    frequency_labels = {
        "semanal": "Semanal",
        "quincenal": "Quincenal",
        "mensual": "Mensual",
        "bimestral": "Bimestral",
        "anual": "Anual",
    }

    try:
        ensure_nails_expenses_table(cur)
        conn.commit()

        today = date.today()
        default_month = today.strftime("%Y-%m")

        if request.method == "POST":
            title = clean_text(request.form.get("title"), 160)
            category = clean_text(request.form.get("category"), 40) or "otros"
            amount = parse_positive_float(request.form.get("amount"), default=0)
            expense_date = parse_date_value(request.form.get("expense_date")) or today
            payment_method = clean_text(request.form.get("payment_method"), 40) or "efectivo"
            is_recurring = request.form.get("is_recurring") == "on"
            recurring_frequency = clean_text(request.form.get("recurring_frequency"), 40) or None
            recurring_day = parse_positive_int(request.form.get("recurring_day"), default=0, max_value=31) or None
            notes = clean_text(request.form.get("notes"), 1000)

            if not title:
                flash("El nombre del gasto es obligatorio.", "warning")
                return redirect(url_for("nails.gastos"))

            if category not in EXPENSE_CATEGORIES:
                flash("Categoría de gasto inválida.", "warning")
                return redirect(url_for("nails.gastos"))

            if payment_method not in PAYMENT_METHODS:
                flash("Método de pago inválido.", "warning")
                return redirect(url_for("nails.gastos"))

            if is_recurring:
                if recurring_frequency not in EXPENSE_FREQUENCIES:
                    flash("Selecciona una frecuencia válida para el gasto recurrente.", "warning")
                    return redirect(url_for("nails.gastos"))
                recurring_day = recurring_day or min(expense_date.day, 31)
            else:
                recurring_frequency = None
                recurring_day = None

            cur.execute(
                """
                INSERT INTO nails_expenses (
                    business_id, title, category, amount, expense_date,
                    payment_method, is_recurring, recurring_day,
                    recurring_frequency, notes
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    business["id"], title, category, amount, expense_date,
                    payment_method, is_recurring, recurring_day,
                    recurring_frequency, notes,
                ),
            )
            conn.commit()
            flash("Gasto registrado correctamente.", "success")
            return redirect(url_for("nails.gastos", month=expense_date.strftime("%Y-%m")))

        selected_month = clean_text(request.args.get("month"), 7) or default_month
        if not re.fullmatch(r"\d{4}-\d{2}", selected_month):
            selected_month = default_month

        month_start = datetime.strptime(f"{selected_month}-01", "%Y-%m-%d").date()
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1, day=1) - timedelta(days=1)

        q = clean_text(request.args.get("q"), 80)
        filter_status = clean_text(request.args.get("status"), 30) or "todos"
        if filter_status not in {"todos", "activos", "recurrentes", "cancelados"}:
            filter_status = "todos"

        page = max(request.args.get("page", 1, type=int), 1)
        per_page = 10
        offset = (page - 1) * per_page

        where_sql = """
            WHERE business_id = %s
              AND expense_date BETWEEN %s::date AND %s::date
        """
        params = [business["id"], month_start, month_end]

        if q:
            like = f"%{q}%"
            where_sql += """
              AND (
                TRANSLATE(LOWER(title), 'áéíóú', 'aeiou') ILIKE TRANSLATE(LOWER(%s), 'áéíóú', 'aeiou')
                OR COALESCE(notes, '') ILIKE %s
                OR category ILIKE %s
              )
            """
            params.extend([like, like, like])

        if filter_status == "activos":
            where_sql += " AND status = 'activo'"
        elif filter_status == "recurrentes":
            where_sql += " AND status = 'activo' AND is_recurring = TRUE"
        elif filter_status == "cancelados":
            where_sql += " AND status = 'cancelado'"

        cur.execute(f"SELECT COUNT(*) AS total FROM nails_expenses {where_sql}", params)
        total_filtered = cur.fetchone()["total"] or 0
        total_pages = max(1, (total_filtered + per_page - 1) // per_page)

        cur.execute(
            f"""
            SELECT
                *,
                TO_CHAR(expense_date, 'DD/MM/YYYY') AS expense_date_label
            FROM nails_expenses
            {where_sql}
            ORDER BY expense_date DESC, id DESC
            LIMIT %s OFFSET %s
            """,
            [*params, per_page, offset],
        )
        expenses = cur.fetchall()

        cur.execute(
            """
            SELECT
                COALESCE(SUM(amount) FILTER (WHERE status = 'activo'), 0) AS month_total,
                COALESCE(SUM(amount) FILTER (WHERE status = 'activo' AND category = 'materiales'), 0) AS materials_total,
                COUNT(*) FILTER (WHERE status = 'activo' AND is_recurring = TRUE) AS recurring_count,
                COUNT(*) FILTER (WHERE status = 'cancelado') AS cancelled_count
            FROM nails_expenses
            WHERE business_id = %s
              AND expense_date BETWEEN %s::date AND %s::date
            """,
            (business["id"], month_start, month_end),
        )
        stats_row = cur.fetchone()

        cur.execute(
            """
            SELECT
                *,
                TO_CHAR(expense_date, 'DD/MM/YYYY') AS expense_date_label
            FROM nails_expenses
            WHERE business_id = %s
              AND status = 'activo'
              AND is_recurring = TRUE
            ORDER BY
                COALESCE(recurring_day, EXTRACT(DAY FROM expense_date)::INT) ASC,
                title ASC
            LIMIT 6
            """,
            (business["id"],),
        )
        recurring_expenses = cur.fetchall()

        cur.execute(
            """
            SELECT
                *,
                TO_CHAR(expense_date, 'DD/MM/YYYY') AS expense_date_label
            FROM nails_expenses
            WHERE business_id = %s
              AND status = 'activo'
              AND expense_date >= CURRENT_DATE
            ORDER BY expense_date ASC, id ASC
            LIMIT 1
            """,
            (business["id"],),
        )
        next_expense = cur.fetchone()

        stats = {
            "month_total": stats_row["month_total"] or 0,
            "materials_total": stats_row["materials_total"] or 0,
            "recurring_count": stats_row["recurring_count"] or 0,
            "cancelled_count": stats_row["cancelled_count"] or 0,
        }

        return render_template(
            "nails/gastos.html",
            business=business,
            expenses=expenses,
            recurring_expenses=recurring_expenses,
            next_expense=next_expense,
            stats=stats,
            q=q,
            filter_status=filter_status,
            selected_month=selected_month,
            current_date=today.strftime("%Y-%m-%d"),
            month_start=month_start,
            month_end=month_end,
            category_labels=category_labels,
            payment_labels=payment_labels,
            frequency_labels=frequency_labels,
            page=page,
            per_page=per_page,
            total_filtered=total_filtered,
            total_pages=total_pages,
        )

    except Exception as e:
        conn.rollback()
        flash(f"Ocurrió un error en gastos: {e}", "danger")
        return redirect(url_for("nails.dashboard"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/gastos/<int:expense_id>/editar", methods=["POST"])
def editar_gasto(expense_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        ensure_nails_expenses_table(cur)

        title = clean_text(request.form.get("title"), 160)
        category = clean_text(request.form.get("category"), 40) or "otros"
        amount = parse_positive_float(request.form.get("amount"), default=0)
        expense_date = parse_date_value(request.form.get("expense_date")) or date.today()
        payment_method = clean_text(request.form.get("payment_method"), 40) or "efectivo"
        is_recurring = request.form.get("is_recurring") == "on"
        recurring_frequency = clean_text(request.form.get("recurring_frequency"), 40) or None
        recurring_day = parse_positive_int(request.form.get("recurring_day"), default=0, max_value=31) or None
        status = clean_text(request.form.get("status"), 20) or "activo"
        notes = clean_text(request.form.get("notes"), 1000)

        if not title:
            flash("El nombre del gasto es obligatorio.", "warning")
            return redirect(url_for("nails.gastos"))

        if category not in EXPENSE_CATEGORIES or payment_method not in PAYMENT_METHODS or status not in EXPENSE_STATUSES:
            flash("Revisa los datos del gasto.", "warning")
            return redirect(url_for("nails.gastos"))

        if is_recurring:
            if recurring_frequency not in EXPENSE_FREQUENCIES:
                flash("Selecciona una frecuencia válida para el gasto recurrente.", "warning")
                return redirect(url_for("nails.gastos"))
            recurring_day = recurring_day or min(expense_date.day, 31)
        else:
            recurring_frequency = None
            recurring_day = None

        cur.execute(
            """
            UPDATE nails_expenses
            SET title = %s,
                category = %s,
                amount = %s,
                expense_date = %s,
                payment_method = %s,
                is_recurring = %s,
                recurring_day = %s,
                recurring_frequency = %s,
                notes = %s,
                status = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (
                title, category, amount, expense_date, payment_method,
                is_recurring, recurring_day, recurring_frequency, notes,
                status, expense_id, business["id"],
            ),
        )

        if cur.rowcount == 0:
            flash("El gasto no existe o no pertenece a este salón.", "warning")
            conn.rollback()
            return redirect(url_for("nails.gastos"))

        conn.commit()
        flash("Gasto actualizado correctamente.", "success")
        return redirect(url_for("nails.gastos", month=expense_date.strftime("%Y-%m")))

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo editar el gasto: {e}", "danger")
        return redirect(url_for("nails.gastos"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/gastos/<int:expense_id>/cancelar", methods=["POST"])
def cancelar_gasto(expense_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        ensure_nails_expenses_table(cur)
        cur.execute(
            """
            UPDATE nails_expenses
            SET status = 'cancelado',
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (expense_id, business["id"]),
        )

        if cur.rowcount == 0:
            flash("El gasto no existe o no pertenece a este salón.", "warning")
            conn.rollback()
            return redirect(url_for("nails.gastos"))

        conn.commit()
        flash("Gasto cancelado correctamente.", "success")
        return redirect(url_for("nails.gastos"))

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo cancelar el gasto: {e}", "danger")
        return redirect(url_for("nails.gastos"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# REPORTES
# Muestra métricas del periodo seleccionado:
# resumen de ventas, resumen de citas, ventas por estado,
# ventas por día, servicios/extras más vendidos,
# clientas frecuentes y ventas con saldo pendiente.
# =========================================================

@nails_bp.route("/reportes")
def reportes():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        business_timezone = business["timezone"] or "America/Monterrey"
        ensure_nails_expenses_table(cur)
        conn.commit()

        # Rango de fechas por defecto: mes actual
        today        = date.today()
        default_from = today.replace(day=1)
        date_from    = request.args.get("date_from") or default_from.strftime("%Y-%m-%d")
        date_to      = request.args.get("date_to")   or today.strftime("%Y-%m-%d")

        date_from_obj = parse_date_value(date_from) or default_from
        date_to_obj   = parse_date_value(date_to)   or today
        date_from     = date_from_obj.strftime("%Y-%m-%d")
        date_to       = date_to_obj.strftime("%Y-%m-%d")

        # Si el usuario intercambió las fechas, ordenarlas
        if date_from_obj > date_to_obj:
            date_from_obj, date_to_obj = date_to_obj, date_from_obj
            date_from = date_from_obj.strftime("%Y-%m-%d")
            date_to   = date_to_obj.strftime("%Y-%m-%d")

        # ── Resumen de ventas del periodo ──────────────────
        cur.execute(
            """
            SELECT
                COUNT(*)                         AS sales_count,
                COALESCE(SUM(total), 0)          AS total_sales,
                COALESCE(SUM(paid_amount), 0)    AS total_paid,
                COALESCE(SUM(balance_due), 0)    AS total_balance,
                COALESCE(SUM(discount_amount),0) AS total_discount
            FROM nails_sales
            WHERE business_id = %s
              AND status != 'cancelada'
              AND DATE(created_at AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            """,
            (business["id"], business_timezone, date_from, date_to),
        )
        sales_summary = cur.fetchone()

        # ── Gastos del periodo ─────────────────────────────
        cur.execute(
            """
            SELECT
                COUNT(*) AS expenses_count,
                COALESCE(SUM(amount), 0) AS total_expenses
            FROM nails_expenses
            WHERE business_id = %s
              AND status = 'activo'
              AND expense_date BETWEEN %s::date AND %s::date
            """,
            (business["id"], date_from, date_to),
        )
        expenses_summary = cur.fetchone()

        # ── Gastos por categoría ───────────────────────────
        cur.execute(
            """
            SELECT
                category,
                COUNT(*) AS expenses_count,
                COALESCE(SUM(amount), 0) AS total
            FROM nails_expenses
            WHERE business_id = %s
              AND status = 'activo'
              AND expense_date BETWEEN %s::date AND %s::date
            GROUP BY category
            ORDER BY total DESC, expenses_count DESC
            """,
            (business["id"], date_from, date_to),
        )
        expenses_by_category_raw = cur.fetchall()

        # ── Resumen de citas del periodo ───────────────────
        cur.execute(
            """
            SELECT
                COUNT(*) AS appointments_count,
                COUNT(*) FILTER (WHERE status = 'pendiente')   AS pending_count,
                COUNT(*) FILTER (WHERE status = 'confirmada')  AS confirmed_count,
                COUNT(*) FILTER (WHERE status = 'atendida')    AS attended_count,
                COUNT(*) FILTER (WHERE status = 'cancelada')   AS cancelled_count,
                COUNT(*) FILTER (WHERE status = 'no_asistio')  AS no_show_count
            FROM nails_appointments
            WHERE business_id = %s
              AND DATE(start_time AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            """,
            (business["id"], business_timezone, date_from, date_to),
        )
        appointments_summary = cur.fetchone()

        # ── Ventas agrupadas por estado ────────────────────
        cur.execute(
            """
            SELECT
                status,
                COUNT(*)               AS count,
                COALESCE(SUM(total),0) AS total
            FROM nails_sales
            WHERE business_id = %s
              AND DATE(created_at AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            GROUP BY status
            ORDER BY total DESC
            """,
            (business["id"], business_timezone, date_from, date_to),
        )
        sales_by_status = cur.fetchall()

        # ── Ventas agrupadas por día ───────────────────────
        cur.execute(
            """
            SELECT
                DATE(created_at AT TIME ZONE %s) AS sale_date,
                COUNT(*)                          AS sales_count,
                COALESCE(SUM(total), 0)           AS total_sales,
                COALESCE(SUM(paid_amount), 0)     AS total_paid
            FROM nails_sales
            WHERE business_id = %s
              AND status != 'cancelada'
              AND DATE(created_at AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            GROUP BY DATE(created_at AT TIME ZONE %s)
            ORDER BY sale_date ASC
            """,
            (business_timezone, business["id"],
             business_timezone, date_from, date_to, business_timezone),
        )
        sales_by_day = cur.fetchall()

        # ── Top 10 servicios más vendidos ──────────────────
        cur.execute(
            """
            SELECT
                d.name,
                COALESCE(SUM(d.quantity), 0) AS quantity,
                COALESCE(SUM(d.total),    0) AS total
            FROM nails_sale_details d
            INNER JOIN nails_sales v ON v.id = d.sale_id
            WHERE v.business_id = %s
              AND v.status != 'cancelada'
              AND d.item_type = 'service'
              AND DATE(v.created_at AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            GROUP BY d.name
            ORDER BY quantity DESC, total DESC
            LIMIT 10
            """,
            (business["id"], business_timezone, date_from, date_to),
        )
        top_services = cur.fetchall()

        # ── Top 10 extras más vendidos ─────────────────────
        cur.execute(
            """
            SELECT
                d.name,
                COALESCE(SUM(d.quantity), 0) AS quantity,
                COALESCE(SUM(d.total),    0) AS total
            FROM nails_sale_details d
            INNER JOIN nails_sales v ON v.id = d.sale_id
            WHERE v.business_id = %s
              AND v.status != 'cancelada'
              AND d.item_type = 'extra'
              AND DATE(v.created_at AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            GROUP BY d.name
            ORDER BY quantity DESC, total DESC
            LIMIT 10
            """,
            (business["id"], business_timezone, date_from, date_to),
        )
        top_extras = cur.fetchall()

        # ── Top 10 clientas por gasto total ───────────────
        cur.execute(
            """
            SELECT
                COALESCE(c.name, 'Cliente General') AS client_name,
                c.phone                              AS client_phone,
                COUNT(v.id)                          AS sales_count,
                COALESCE(SUM(v.total),       0)      AS total_spent,
                COALESCE(SUM(v.balance_due), 0)      AS total_balance
            FROM nails_sales v
            LEFT JOIN nails_clients c ON c.id = v.client_id
            WHERE v.business_id = %s
              AND v.status != 'cancelada'
              AND DATE(v.created_at AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            GROUP BY c.name, c.phone
            ORDER BY total_spent DESC, sales_count DESC
            LIMIT 10
            """,
            (business["id"], business_timezone, date_from, date_to),
        )
        top_clients = cur.fetchall()

        # ── Ventas con saldo pendiente (sin filtro de fechas) ─
        cur.execute(
            """
            SELECT
                v.*,
                c.name  AS client_name,
                c.phone AS client_phone,
                TO_CHAR(v.created_at AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS created_at_formatted
            FROM nails_sales v
            LEFT JOIN nails_clients c ON c.id = v.client_id
            WHERE v.business_id = %s
              AND v.status IN ('pendiente', 'anticipo')
              AND v.balance_due > 0
            ORDER BY v.created_at DESC
            LIMIT 10
            """,
            (business_timezone, business["id"]),
        )
        pending_sales = cur.fetchall()

        status_labels = {
            "pagada": "Pagada",
            "anticipo": "Anticipo",
            "pendiente": "Pendiente",
            "cancelada": "Cancelada",
        }
        status_map = {row["status"]: row for row in sales_by_status}
        status_total = sum(float(row["total"] or 0) for row in sales_by_status) or 0
        sales_status_rows = []
        for status_key in ("pagada", "anticipo", "pendiente", "cancelada"):
            row = status_map.get(status_key, {})
            amount = float(row.get("total", 0) or 0)
            count = int(row.get("count", 0) or 0)
            percent = round((amount / status_total) * 100) if status_total else 0
            sales_status_rows.append({
                "status": status_key,
                "label": status_labels[status_key],
                "count": count,
                "total": amount,
                "percent": percent,
            })

        max_day_total = max([float(day["total_sales"] or 0) for day in sales_by_day] or [0])
        expense_category_labels = {
            "materiales": "Materiales",
            "renta": "Renta",
            "servicios": "Servicios",
            "sueldos": "Sueldos",
            "comisiones": "Comisiones",
            "publicidad": "Publicidad",
            "mantenimiento": "Mantenimiento",
            "capacitacion": "Capacitación",
            "otros": "Otros",
        }
        total_expenses_value = float(expenses_summary["total_expenses"] or 0)
        expenses_by_category = []
        for row in expenses_by_category_raw:
            amount = float(row["total"] or 0)
            expenses_by_category.append({
                "category": row["category"],
                "label": expense_category_labels.get(row["category"], (row["category"] or "Otros").title()),
                "expenses_count": row["expenses_count"] or 0,
                "total": amount,
                "percent": round((amount / total_expenses_value) * 100) if total_expenses_value else 0,
            })

        today_link = today.strftime("%Y-%m-%d")
        quick_ranges = {
            "today": {
                "date_from": today_link,
                "date_to": today_link,
            },
            "seven": {
                "date_from": (today - timedelta(days=6)).strftime("%Y-%m-%d"),
                "date_to": today_link,
            },
            "thirty": {
                "date_from": (today - timedelta(days=29)).strftime("%Y-%m-%d"),
                "date_to": today_link,
            },
        }

        total_sales_value = float(sales_summary["total_sales"] or 0)
        total_paid_value = float(sales_summary["total_paid"] or 0)
        sales_count_value = int(sales_summary["sales_count"] or 0)
        net_profit_value = total_sales_value - total_expenses_value
        appointment_count_value = int(appointments_summary["appointments_count"] or 0)
        cancelled_count_value = int(appointments_summary["cancelled_count"] or 0)
        no_show_count_value = int(appointments_summary["no_show_count"] or 0)
        attended_count_value = int(appointments_summary["attended_count"] or 0)
        effective_appointments = max(appointment_count_value - cancelled_count_value, 0)
        biggest_expense = expenses_by_category[0] if expenses_by_category else None

        stats = {
            "sales_count":        sales_summary["sales_count"]          or 0,
            "total_sales":        sales_summary["total_sales"]           or 0,
            "total_paid":         sales_summary["total_paid"]            or 0,
            "total_balance":      sales_summary["total_balance"]         or 0,
            "total_discount":     sales_summary["total_discount"]        or 0,
            "expenses_count":     expenses_summary["expenses_count"]     or 0,
            "total_expenses":     expenses_summary["total_expenses"]      or 0,
            "net_profit":         net_profit_value,
            "avg_ticket":         (total_sales_value / sales_count_value) if sales_count_value else 0,
            "expense_ratio":      round((total_expenses_value / total_sales_value) * 100) if total_sales_value else 0,
            "net_margin":         round((net_profit_value / total_sales_value) * 100) if total_sales_value else 0,
            "paid_ratio":         round((total_paid_value / total_sales_value) * 100) if total_sales_value else 0,
            "attendance_rate":    round((attended_count_value / effective_appointments) * 100) if effective_appointments else 0,
            "no_show_rate":       round((no_show_count_value / effective_appointments) * 100) if effective_appointments else 0,
            "biggest_expense_label": biggest_expense["label"] if biggest_expense else "Sin gastos",
            "biggest_expense_total": biggest_expense["total"] if biggest_expense else 0,
            "biggest_expense_percent": biggest_expense["percent"] if biggest_expense else 0,
            "appointments_count": appointments_summary["appointments_count"] or 0,
            "pending_count":      appointments_summary["pending_count"]      or 0,
            "confirmed_count":    appointments_summary["confirmed_count"]    or 0,
            "attended_count":     appointments_summary["attended_count"]     or 0,
            "cancelled_count":    appointments_summary["cancelled_count"]    or 0,
            "no_show_count":      appointments_summary["no_show_count"]      or 0,
        }

        return render_template(
            "nails/reportes.html",
            business=business,
            stats=stats,
            date_from=date_from,
            date_to=date_to,
            sales_by_status=sales_by_status,
            sales_status_rows=sales_status_rows,
            sales_by_day=sales_by_day,
            max_day_total=max_day_total,
            top_services=top_services,
            top_extras=top_extras,
            top_clients=top_clients,
            expenses_by_category=expenses_by_category,
            pending_sales=pending_sales,
            quick_ranges=quick_ranges,
        )

    except Exception as e:
        conn.rollback()
        flash(f"No se pudieron cargar los reportes: {e}", "danger")
        return redirect(url_for("nails.dashboard"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/reportes/excel")
def reportes_excel():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        business_timezone = business["timezone"] or "America/Monterrey"
        ensure_nails_expenses_table(cur)
        conn.commit()

        today        = date.today()
        default_from = today.replace(day=1)
        date_from    = request.args.get("date_from") or default_from.strftime("%Y-%m-%d")
        date_to      = request.args.get("date_to")   or today.strftime("%Y-%m-%d")
        date_from_obj = parse_date_value(date_from) or default_from
        date_to_obj   = parse_date_value(date_to)   or today

        if date_from_obj > date_to_obj:
            date_from_obj, date_to_obj = date_to_obj, date_from_obj

        date_from = date_from_obj.strftime("%Y-%m-%d")
        date_to   = date_to_obj.strftime("%Y-%m-%d")

        cur.execute(
            """
            SELECT
                COUNT(*) AS ventas,
                COALESCE(SUM(total), 0) AS total_ventas,
                COALESCE(SUM(paid_amount), 0) AS cobrado,
                COALESCE(SUM(balance_due), 0) AS saldo,
                COALESCE(SUM(discount_amount), 0) AS descuentos
            FROM nails_sales
            WHERE business_id = %s
              AND status != 'cancelada'
              AND DATE(created_at AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            """,
            (business["id"], business_timezone, date_from, date_to),
        )
        sales_summary = cur.fetchone()

        cur.execute(
            """
            SELECT
                COUNT(*) AS gastos,
                COALESCE(SUM(amount), 0) AS total_gastos
            FROM nails_expenses
            WHERE business_id = %s
              AND status = 'activo'
              AND expense_date BETWEEN %s::date AND %s::date
            """,
            (business["id"], date_from, date_to),
        )
        expenses_summary = cur.fetchone()

        cur.execute(
            """
            SELECT
                v.sale_number AS folio,
                TO_CHAR(v.created_at AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS fecha,
                COALESCE(c.name, 'Cliente General') AS clienta,
                c.phone AS telefono,
                COALESCE(st.name, 'Sin asignar') AS tecnica,
                v.status AS estado,
                v.payment_method AS metodo_pago,
                v.subtotal,
                v.discount_amount AS descuento,
                v.tax_amount AS impuestos,
                v.total,
                v.paid_amount AS pagado,
                v.balance_due AS saldo,
                v.notes AS notas
            FROM nails_sales v
            LEFT JOIN nails_clients c ON c.id = v.client_id
            LEFT JOIN nails_staff st ON st.id = v.staff_id
            WHERE v.business_id = %s
              AND DATE(v.created_at AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            ORDER BY v.created_at DESC
            """,
            (business_timezone, business["id"], business_timezone, date_from, date_to),
        )
        sales_rows = cur.fetchall()

        cur.execute(
            """
            SELECT
                v.sale_number AS folio,
                TO_CHAR(v.created_at AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS fecha,
                COALESCE(c.name, 'Cliente General') AS clienta,
                d.item_type AS tipo,
                d.name AS concepto,
                d.description AS descripcion,
                d.quantity AS cantidad,
                d.unit_price AS precio_unitario,
                d.total AS total
            FROM nails_sale_details d
            INNER JOIN nails_sales v ON v.id = d.sale_id
            LEFT JOIN nails_clients c ON c.id = v.client_id
            WHERE v.business_id = %s
              AND v.status != 'cancelada'
              AND DATE(v.created_at AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            ORDER BY v.created_at DESC, d.id ASC
            """,
            (business_timezone, business["id"], business_timezone, date_from, date_to),
        )
        sale_details_rows = cur.fetchall()

        cur.execute(
            """
            SELECT
                expense_date AS fecha,
                title AS gasto,
                category AS categoria,
                amount AS monto,
                payment_method AS metodo_pago,
                is_recurring AS recurrente,
                recurring_frequency AS frecuencia,
                recurring_day AS dia_recurrente,
                status AS estado,
                notes AS notas
            FROM nails_expenses
            WHERE business_id = %s
              AND expense_date BETWEEN %s::date AND %s::date
            ORDER BY expense_date DESC, id DESC
            """,
            (business["id"], date_from, date_to),
        )
        expense_rows = cur.fetchall()

        cur.execute(
            """
            SELECT
                category AS categoria,
                COUNT(*) AS registros,
                COALESCE(SUM(amount), 0) AS total
            FROM nails_expenses
            WHERE business_id = %s
              AND status = 'activo'
              AND expense_date BETWEEN %s::date AND %s::date
            GROUP BY category
            ORDER BY total DESC
            """,
            (business["id"], date_from, date_to),
        )
        expense_category_rows = cur.fetchall()

        cur.execute(
            """
            SELECT
                TO_CHAR(a.start_time AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS inicio,
                TO_CHAR(a.end_time AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS fin,
                COALESCE(c.name, 'Cliente General') AS clienta,
                c.phone AS telefono,
                COALESCE(st.name, 'Sin asignar') AS tecnica,
                a.status AS estado,
                COALESCE(
                    STRING_AGG(DISTINCT aps.name, ', ' ORDER BY aps.name),
                    a.title,
                    s.name,
                    'Sin servicio'
                ) AS servicios,
                a.estimated_total AS total_estimado,
                a.deposit_amount AS anticipo,
                a.notes AS notas
            FROM nails_appointments a
            LEFT JOIN nails_clients c ON c.id = a.client_id
            LEFT JOIN nails_staff st ON st.id = a.staff_id
            LEFT JOIN nails_services s ON s.id = a.service_id
            LEFT JOIN nails_appointment_services aps ON aps.appointment_id = a.id
            WHERE a.business_id = %s
              AND DATE(a.start_time AT TIME ZONE %s) BETWEEN %s::date AND %s::date
            GROUP BY a.id, c.name, c.phone, st.name, s.name
            ORDER BY a.start_time ASC
            """,
            (business_timezone, business_timezone, business["id"], business_timezone, date_from, date_to),
        )
        appointment_rows = cur.fetchall()

        wb = Workbook()
        wb.remove(wb.active)
        header_fill = PatternFill("solid", fgColor="8F4AA0")
        header_font = Font(color="FFFFFF", bold=True)
        soft_fill = PatternFill("solid", fgColor="F8E7F7")
        border = Border(bottom=Side(style="thin", color="E7D9E9"))

        def money_value(value):
            return float(value or 0)

        def add_sheet(title, headers, rows):
            ws = wb.create_sheet(title)
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row in rows:
                ws.append([row.get(header, "") for header in headers])
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 36)
            ws.freeze_panes = "A2"

        total_sales = money_value(sales_summary["total_ventas"])
        total_expenses = money_value(expenses_summary["total_gastos"])
        net_profit = total_sales - total_expenses
        sales_count = int(sales_summary["ventas"] or 0)
        avg_ticket = total_sales / sales_count if sales_count else 0

        summary_ws = wb.create_sheet("Resumen")
        summary_ws.append(["Salón", business["name"] or "Sianeffects Nails"])
        summary_ws.append(["Periodo", f"{date_from} a {date_to}"])
        summary_ws.append([])
        summary_ws.append(["Indicador", "Valor"])
        summary_ws.append(["Ventas", sales_count])
        summary_ws.append(["Total vendido", total_sales])
        summary_ws.append(["Cobrado", money_value(sales_summary["cobrado"])])
        summary_ws.append(["Saldo pendiente", money_value(sales_summary["saldo"])])
        summary_ws.append(["Descuentos", money_value(sales_summary["descuentos"])])
        summary_ws.append(["Gastos", total_expenses])
        summary_ws.append(["Utilidad neta", net_profit])
        summary_ws.append(["Ticket promedio", avg_ticket])
        for row in summary_ws.iter_rows():
            for cell in row:
                cell.border = border
        for cell in summary_ws[4]:
            cell.fill = header_fill
            cell.font = header_font
        summary_ws["A1"].font = Font(bold=True, size=14)
        summary_ws["B1"].fill = soft_fill
        summary_ws.column_dimensions["A"].width = 22
        summary_ws.column_dimensions["B"].width = 30

        add_sheet("Ventas", [
            "folio", "fecha", "clienta", "telefono", "tecnica", "estado", "metodo_pago",
            "subtotal", "descuento", "impuestos", "total", "pagado", "saldo", "notas",
        ], [dict(row) for row in sales_rows])
        add_sheet("Detalle ventas", [
            "folio", "fecha", "clienta", "tipo", "concepto", "descripcion",
            "cantidad", "precio_unitario", "total",
        ], [dict(row) for row in sale_details_rows])
        add_sheet("Gastos", [
            "fecha", "gasto", "categoria", "monto", "metodo_pago", "recurrente",
            "frecuencia", "dia_recurrente", "estado", "notas",
        ], [dict(row) for row in expense_rows])
        add_sheet("Gastos por categoría", [
            "categoria", "registros", "total",
        ], [dict(row) for row in expense_category_rows])
        add_sheet("Citas", [
            "inicio", "fin", "clienta", "telefono", "tecnica", "estado",
            "servicios", "total_estimado", "anticipo", "notas",
        ], [dict(row) for row in appointment_rows])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        safe_slug = re.sub(r"[^a-zA-Z0-9_-]+", "-", business.get("slug") or business.get("name") or "nails").strip("-").lower()
        filename = f"reporte_nails_{safe_slug}_{date_from}_a_{date_to}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo generar el Excel: {e}", "danger")
        return redirect(url_for("nails.reportes", date_from=request.args.get("date_from"), date_to=request.args.get("date_to")))

    finally:
        cur.close()
        conn.close()


# =========================================================
# PERSONAL
# Jefa administra técnicas y roles del salón.
# =========================================================

@nails_bp.route("/personal", methods=["GET"])
def personal():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        is_owner = user_is_nails_owner(cur, user_id, business)

        cur.execute(
            """
            SELECT
                st.*,
                u.username,
                u.email AS user_email,
                COALESCE(a.appointments_count, 0) AS appointments_count,
                COALESCE(v.sales_total, 0) AS sales_total
            FROM nails_staff st
            LEFT JOIN usuarios u ON u.id = st.user_id
            LEFT JOIN (
                SELECT staff_id, COUNT(*) AS appointments_count
                FROM nails_appointments
                WHERE business_id = %s
                  AND staff_id IS NOT NULL
                  AND status NOT IN ('cancelada', 'no_asistio')
                GROUP BY staff_id
            ) a ON a.staff_id = st.id
            LEFT JOIN (
                SELECT staff_id, COALESCE(SUM(total), 0) AS sales_total
                FROM nails_sales
                WHERE business_id = %s
                  AND staff_id IS NOT NULL
                  AND status != 'cancelada'
                GROUP BY staff_id
            ) v ON v.staff_id = st.id
            WHERE st.business_id = %s
            ORDER BY
                CASE st.role WHEN 'owner' THEN 1 WHEN 'admin' THEN 2 WHEN 'reception' THEN 3 ELSE 4 END,
                st.is_active DESC,
                LOWER(st.name) ASC
            """,
            (business["id"], business["id"], business["id"]),
        )
        staff_members = cur.fetchall()

        return render_template(
            "nails/personal.html",
            business=business,
            staff_members=staff_members,
            is_owner=is_owner,
            role_labels=NAILS_ROLE_LABELS,
            role_descriptions=NAILS_ROLE_DESCRIPTIONS,
            current_user_id=user_id,
        )

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo cargar personal: {e}", "danger")
        return redirect(url_for("nails.dashboard"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/personal/<int:staff_id>/editar", methods=["POST"])
def editar_personal(staff_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        if not user_is_nails_owner(cur, user_id, business):
            flash("Solo la jefa puede editar personal.", "warning")
            return redirect(url_for("nails.personal"))

        name = clean_text(request.form.get("name"), 120)
        email = clean_text(request.form.get("email"), 120).lower()
        phone = clean_text(request.form.get("phone"), 40)
        role = clean_text(request.form.get("role"), 40) or "staff"
        color = clean_hex_color(request.form.get("color"), DEFAULT_PRIMARY_COLOR)
        commission_type = "none"
        commission_value = 0
        is_active = request.form.get("is_active") == "on"

        if not name or role not in NAILS_STAFF_ROLES:
            flash("Revisa nombre y rol.", "warning")
            return redirect(url_for("nails.personal"))

        cur.execute(
            """
            SELECT role, user_id
            FROM nails_staff
            WHERE id = %s AND business_id = %s
            LIMIT 1
            """,
            (staff_id, business["id"]),
        )
        current_staff = cur.fetchone()
        if not current_staff:
            flash("No se encontró ese miembro del personal.", "warning")
            return redirect(url_for("nails.personal"))

        editing_self = current_staff["user_id"] and int(current_staff["user_id"]) == int(user_id)
        if editing_self and current_staff["role"] == "owner" and role != "owner":
            flash("No puedes quitarte tu propio rol de jefa.", "warning")
            return redirect(url_for("nails.personal"))

        if editing_self and not is_active:
            flash("No puedes desactivarte a ti misma.", "warning")
            return redirect(url_for("nails.personal"))

        if current_staff["role"] == "owner" and role != "owner":
            cur.execute(
                """
                SELECT COUNT(*) AS owners_count
                FROM nails_staff
                WHERE business_id = %s AND role = 'owner' AND is_active = TRUE
                """,
                (business["id"],),
            )
            if (cur.fetchone()["owners_count"] or 0) <= 1:
                flash("Debe existir al menos una jefa activa.", "warning")
                return redirect(url_for("nails.personal"))

        cur.execute(
            """
            UPDATE nails_staff
            SET name = %s,
                email = %s,
                phone = %s,
                role = %s,
                color = %s,
                commission_type = %s,
                commission_value = %s,
                is_active = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (
                name, email, phone, role, color, commission_type,
                commission_value, is_active, staff_id, business["id"],
            ),
        )
        conn.commit()
        flash("Personal actualizado correctamente.", "success")
        return redirect(url_for("nails.personal"))

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo editar personal: {e}", "danger")
        return redirect(url_for("nails.personal"))

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/personal/<int:staff_id>/desactivar", methods=["POST"])
def desactivar_personal(staff_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        if not user_is_nails_owner(cur, user_id, business):
            flash("Solo la jefa puede desactivar personal.", "warning")
            return redirect(url_for("nails.personal"))

        cur.execute(
            """
            SELECT role, user_id
            FROM nails_staff
            WHERE id = %s AND business_id = %s
            LIMIT 1
            """,
            (staff_id, business["id"]),
        )
        staff = cur.fetchone()
        if not staff:
            flash("No se encontró ese miembro del personal.", "warning")
            return redirect(url_for("nails.personal"))

        if staff["user_id"] and int(staff["user_id"]) == int(user_id):
            flash("No puedes desactivarte a ti misma.", "warning")
            return redirect(url_for("nails.personal"))

        if staff["role"] == "owner":
            cur.execute(
                """
                SELECT COUNT(*) AS owners_count
                FROM nails_staff
                WHERE business_id = %s AND role = 'owner' AND is_active = TRUE
                """,
                (business["id"],),
            )
            if (cur.fetchone()["owners_count"] or 0) <= 1:
                flash("No puedes desactivar a la única jefa activa.", "warning")
                return redirect(url_for("nails.personal"))

        cur.execute(
            """
            UPDATE nails_staff
            SET is_active = FALSE,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (staff_id, business["id"]),
        )
        conn.commit()
        flash("Personal desactivado correctamente.", "success")
        return redirect(url_for("nails.personal"))

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo desactivar personal: {e}", "danger")
        return redirect(url_for("nails.personal"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# CONFIGURACIÓN DEL SALÓN
# GET:  Muestra el formulario con los datos actuales del salón.
# POST: Actualiza nombre, contacto, colores, zona horaria,
#       políticas, horarios y logo (sube a R2 si se adjunta).
# =========================================================

@nails_bp.route("/configuracion", methods=["GET", "POST"])
def configuracion():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        if request.method == "POST":
            name                = clean_text(request.form.get("name"), 120)
            whatsapp            = clean_text(request.form.get("whatsapp"), 40)
            instagram           = clean_text(request.form.get("instagram"), 80)
            address             = clean_text(request.form.get("address"), 240)
            logo_url            = business["logo_url"] or ""
            primary_color       = clean_hex_color(request.form.get("primary_color"),   DEFAULT_PRIMARY_COLOR)
            secondary_color     = clean_hex_color(request.form.get("secondary_color"), DEFAULT_SECONDARY_COLOR)
            accent_color        = clean_hex_color(request.form.get("accent_color"),    DEFAULT_ACCENT_COLOR)
            timezone            = clean_text(request.form.get("timezone", "America/Monterrey"), 80)
            currency            = clean_text(request.form.get("currency", "MXN"), 8).upper()
            cancellation_policy = clean_text(request.form.get("cancellation_policy"), 1500)
            deposit_policy      = clean_text(request.form.get("deposit_policy"), 1500)
            catalog_tagline     = clean_text(request.form.get("catalog_tagline"), 240)

            if not name:
                flash("El nombre del salón es obligatorio.", "warning")
                return redirect(url_for("nails.configuracion"))

            # Whitelist de zonas horarias soportadas
            if timezone not in {"America/Monterrey", "America/Mexico_City",
                                 "America/Cancun", "America/Tijuana"}:
                timezone = "America/Monterrey"

            if currency not in {"MXN", "USD"}:
                currency = "MXN"

            try:
                business_hours_json = normalize_business_hours_from_form(request.form)
            except ValueError as e:
                flash(str(e), "warning")
                return redirect(url_for("nails.configuracion"))

            # ── Subir logo si se adjuntó un archivo ────────
            if "logo_file" in request.files:
                file = request.files["logo_file"]

                if file and file.filename:
                    if not allowed_image_file(file.filename):
                        flash("El logo debe ser PNG, JPG, JPEG, SVG o WEBP.", "warning")
                        return redirect(url_for("nails.configuracion"))

                    file.seek(0, os.SEEK_END)
                    file_size = file.tell()
                    file.seek(0)

                    if file_size > 8 * 1024 * 1024:
                        flash("El logo no puede pesar más de 8 MB.", "warning")
                        return redirect(url_for("nails.configuracion"))

                    # Borrar logo anterior de R2 si existe
                    if logo_url and PUBLIC_URL and logo_url.startswith(PUBLIC_URL):
                        try:
                            old_key = logo_url.replace(f"{PUBLIC_URL}/", "")
                            get_s3_client().delete_object(Bucket=BUCKET_NAME, Key=old_key)
                        except Exception as e:
                            current_app.logger.warning(
                                f"NAILS_R2_DELETE_WARNING: No se pudo borrar logo anterior "
                                f"de negocio {business['id']} - {e}"
                            )

                    base_filename   = secure_filename(file.filename)
                    unique_filename = (
                        f"nails/business_{business['id']}/logo/"
                        f"{uuid.uuid4().hex}_{base_filename}"
                    )

                    try:
                        get_s3_client().upload_fileobj(
                            file,
                            BUCKET_NAME,
                            unique_filename,
                            ExtraArgs={"ContentType": file.content_type},
                        )
                        logo_url = f"{PUBLIC_URL}/{unique_filename}"
                        current_app.logger.info(
                            f"NAILS_R2_LOGO_UPLOAD_SUCCESS: Usuario "
                            f"'{session.get('username', 'Anonimo')}' ID {user_id} "
                            f"subió logo '{unique_filename}'"
                        )
                    except Exception as e:
                        current_app.logger.error(
                            f"NAILS_R2_LOGO_UPLOAD_ERROR: Usuario ID {user_id} "
                            f"falló al subir logo - {e}"
                        )
                        flash("Error al subir el logo a la nube.", "danger")
                        return redirect(url_for("nails.configuracion"))

            # ── Guardar cambios en la BD ───────────────────
            cur.execute(
                """
                UPDATE nails_businesses
                SET name                = %s,
                    whatsapp            = %s,
                    instagram           = %s,
                    address             = %s,
                    logo_url            = %s,
                    primary_color       = %s,
                    secondary_color     = %s,
                    accent_color        = %s,
                    timezone            = %s,
                    currency            = %s,
                    cancellation_policy = %s,
                    deposit_policy      = %s,
                    catalog_tagline     = %s,
                    business_hours_json = %s,
                    updated_at          = CURRENT_TIMESTAMP
                WHERE id      = %s
                RETURNING id
                """,
                (
                    name, whatsapp, instagram, address, logo_url,
                    primary_color, secondary_color, accent_color,
                    timezone, currency, cancellation_policy, deposit_policy, catalog_tagline,
                    json.dumps(business_hours_json, ensure_ascii=False),
                    business["id"],
                ),
            )
            updated = cur.fetchone()

            if not updated:
                conn.rollback()
                flash("No se pudo actualizar la configuración.", "warning")
                return redirect(url_for("nails.configuracion"))

            # Registrar en el log de actividad
            cur.execute(
                """
                INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    business["id"], user_id,
                    "business_config_update", "Configuración",
                    "Actualizó la configuración del salón",
                ),
            )
            conn.commit()
            flash("Configuración actualizada correctamente.", "success")
            return redirect(url_for("nails.configuracion"))

        # ── GET: parsear horarios desde JSON almacenado ────
        # BUG CORREGIDO: `import json` era importado dos veces dentro
        # de esta función (una en el bloque POST y otra aquí).
        # Se movió al inicio del módulo.
        business_hours = {}
        if business["business_hours_json"]:
            try:
                business_hours = json.loads(business["business_hours_json"])
            except Exception:
                business_hours = {}

        cur.execute(
            """
            SELECT email
            FROM usuarios
            WHERE id = %s
            LIMIT 1
            """,
            (business["user_id"] or user_id,),
        )
        registered_user = cur.fetchone()
        registered_email = registered_user["email"] if registered_user else ""

        return render_template(
            "nails/configuracion.html",
            business=business,
            business_hours=business_hours,
            business_hours_controls=build_business_hours_controls(business_hours),
            registered_email=registered_email,
            is_owner=user_is_nails_owner(cur, user_id, business),
        )

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo cargar configuración: {e}", "danger")
        return redirect(url_for("nails.dashboard"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# TICKET DE VENTA
# Genera la vista de impresión de un ticket dado su sale_id.
# Convierte el logo a Data URI para que funcione offline/impreso.
# Registra cada visualización en el log de actividad.
# =========================================================

@nails_bp.route("/ventas/<int:sale_id>/ticket")
def ticket(sale_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        business_timezone = business["timezone"] or "America/Monterrey"

        cur.execute(
            """
            SELECT
                v.*,
                c.name  AS client_name,
                c.phone AS client_phone,
                c.email AS client_email,
                st.name AS staff_name,
                TO_CHAR(v.created_at AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS created_at_formatted
            FROM nails_sales v
            LEFT JOIN nails_clients c  ON c.id  = v.client_id
            LEFT JOIN nails_staff   st ON st.id = v.staff_id
            WHERE v.id = %s AND v.business_id = %s
            LIMIT 1
            """,
            (business_timezone, sale_id, business["id"]),
        )
        sale = cur.fetchone()

        if not sale:
            return "Ticket no encontrado", 404

        cur.execute(
            "SELECT * FROM nails_sale_details WHERE sale_id = %s ORDER BY id ASC",
            (sale_id,),
        )
        details = cur.fetchall()

        cur.execute(
            """
            SELECT
                *,
                TO_CHAR(created_at AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS created_at_formatted
            FROM nails_payments
            WHERE sale_id = %s
            ORDER BY created_at ASC
            """,
            (business_timezone, sale_id),
        )
        payments = cur.fetchall()

        current_app.logger.info(
            f"NAILS_TICKET_VIEW: Usuario '{session.get('username', 'Visitante')}' "
            f"visualizó el ticket Nails #{sale_id}"
        )

        # Registrar visualización en log (sin propagar errores al usuario)
        try:
            cur.execute(
                """
                INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    business["id"], user_id,
                    "ticket_view", "Tickets",
                    f"Visualizó el ticket Nails #{sale_id}",
                ),
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            current_app.logger.error(f"Error al registrar log de ticket Nails: {e}")

        # Convertir logo a Data URI; si falla, usa la URL directa como fallback
        ticket_logo_url = image_url_to_data_uri(business["logo_url"]) or business["logo_url"]

        return render_template(
            "nails/ticket.html",
            business=business,
            sale=sale,
            details=details,
            payments=payments,
            ticket_logo_url=ticket_logo_url,
        )

    except Exception as e:
        current_app.logger.error(f"Error al cargar ticket Nails #{sale_id}: {e}")
        return "Error al cargar ticket", 500

    finally:
        cur.close()
        conn.close()


# =========================================================
# CAMBIAR ESTADO DE CITA
# POST: Actualiza el estado de una cita a uno de los 5 posibles.
#       Solo acepta citas que pertenezcan al salón activo.
# =========================================================

@nails_bp.route("/agenda/<int:appointment_id>/estado", methods=["POST"])
def cambiar_estado_cita(appointment_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    nuevo_estado = request.form.get("status", "").strip()

    if nuevo_estado not in APPOINTMENT_STATUSES:
        flash("Estado de cita inválido.", "warning")
        return redirect(url_for("nails.agenda"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            UPDATE nails_appointments
            SET status     = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id          = %s
              AND business_id = %s
            RETURNING id
            """,
            (nuevo_estado, appointment_id, business["id"]),
        )
        updated = cur.fetchone()

        if not updated:
            conn.rollback()
            flash("No se encontró la cita.", "warning")
            return redirect(url_for("nails.agenda"))

        if nuevo_estado == "cancelada":
            cur.execute(
                """
                UPDATE nails_sales
                SET status = 'cancelada',
                    balance_due = 0,
                    updated_at = CURRENT_TIMESTAMP
                WHERE appointment_id = %s
                  AND business_id = %s
                  AND status != 'cancelada'
                """,
                (appointment_id, business["id"]),
            )

        cur.execute(
            """
            INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (
                business["id"], user_id,
                "appointment_status_update", "Agenda",
                f"Cambió la cita #{appointment_id} a estado {nuevo_estado}",
            ),
        )
        conn.commit()
        flash("Estado de cita actualizado.", "success")
        return redirect(url_for("nails.agenda"))

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo actualizar la cita: {e}", "danger")
        return redirect(url_for("nails.agenda"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# EDITAR CITA
# POST: Actualiza datos operativos de una cita y reconstruye
#      servicios/extras para mantener totales y duración consistentes.
# =========================================================

@nails_bp.route("/agenda/<int:appointment_id>/editar", methods=["POST"])
def editar_cita(appointment_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        ensure_nails_appointment_services_table(cur)

        client_id          = clean_optional_id(request.form.get("client_id"))
        staff_id           = clean_optional_id(request.form.get("staff_id"))
        appointment_date   = clean_text(request.form.get("appointment_date"))
        start_time         = clean_text(request.form.get("start_time"))
        status             = clean_text(request.form.get("status", "pendiente"))
        deposit_amount_raw = request.form.get("deposit_amount", "0")
        notes              = clean_text(request.form.get("notes"), 1000)

        selected_service_ids = [
            clean_optional_id(value)
            for value in request.form.getlist("service_ids")
        ]
        selected_service_ids = [value for value in selected_service_ids if value]
        service_id = selected_service_ids[0] if selected_service_ids else None

        selected_extra_ids = [
            clean_optional_id(value)
            for value in request.form.getlist("extras")
        ]
        selected_extra_ids = [value for value in selected_extra_ids if value]

        if not service_id or not appointment_date or not start_time:
            flash("Servicio, fecha y hora son obligatorios.", "warning")
            return redirect(url_for("nails.agenda"))

        if status not in APPOINTMENT_STATUSES:
            flash("Estado de cita inválido.", "warning")
            return redirect(url_for("nails.agenda"))

        if not parse_date_value(appointment_date) or not re.fullmatch(r"\d{2}:\d{2}", start_time):
            flash("Fecha u hora inválida.", "warning")
            return redirect(url_for("nails.agenda"))

        cur.execute(
            """
            SELECT id
            FROM nails_appointments
            WHERE id = %s AND business_id = %s
            LIMIT 1
            """,
            (appointment_id, business["id"]),
        )
        if not cur.fetchone():
            flash("No se encontró la cita.", "warning")
            return redirect(url_for("nails.agenda"))

        if client_id and not row_belongs_to_business(cur, "nails_clients", client_id, business["id"]):
            flash("La clienta seleccionada no pertenece a este salón.", "warning")
            return redirect(url_for("nails.agenda"))

        if staff_id and not row_belongs_to_business(cur, "nails_staff", staff_id, business["id"]):
            flash("La técnica seleccionada no pertenece a este salón.", "warning")
            return redirect(url_for("nails.agenda"))

        unique_service_ids = list(dict.fromkeys(selected_service_ids))

        cur.execute(
            """
            SELECT id, name, base_price, duration_minutes
            FROM nails_services
            WHERE id = ANY(%s)
              AND business_id = %s
              AND is_active = TRUE
            ORDER BY array_position(%s, id)
            """,
            (unique_service_ids, business["id"], unique_service_ids),
        )
        service_rows = cur.fetchall()

        if len(service_rows) != len(unique_service_ids):
            flash("Uno o más servicios seleccionados no existen.", "warning")
            return redirect(url_for("nails.agenda"))

        service_map = {service["id"]: service for service in service_rows}
        services_to_insert = []
        estimated_total = 0
        total_duration = 0

        for index, selected_service_id in enumerate(selected_service_ids):
            service = service_map[selected_service_id]
            service_price = float(service["base_price"] or 0)
            service_duration = int(service["duration_minutes"] or 60)
            estimated_total += service_price
            total_duration += service_duration
            services_to_insert.append({
                "id": service["id"],
                "name": service["name"],
                "price": service_price,
                "duration_minutes": service_duration,
                "sort_order": index,
            })

        extras_to_insert = []

        for extra_id in selected_extra_ids:
            cur.execute(
                """
                SELECT id, name, price, duration_minutes
                FROM nails_extras
                WHERE id = %s
                  AND business_id = %s
                  AND is_active = TRUE
                LIMIT 1
                """,
                (extra_id, business["id"]),
            )
            extra = cur.fetchone()

            if extra:
                extra_price = float(extra["price"] or 0)
                extra_duration = int(extra["duration_minutes"] or 0)
                estimated_total += extra_price
                total_duration += extra_duration
                extras_to_insert.append({
                    "id": extra["id"],
                    "name": extra["name"],
                    "price": extra_price,
                    "duration_minutes": extra_duration,
                })

        business_timezone = business["timezone"] or "America/Monterrey"
        start_datetime_str = f"{appointment_date} {start_time}:00"

        cur.execute(
            """
            SELECT
                (%s::timestamp AT TIME ZONE %s) AS start_time_db,
                ((%s::timestamp AT TIME ZONE %s)
                 + (%s || ' minutes')::interval) AS end_time_db
            """,
            (
                start_datetime_str, business_timezone,
                start_datetime_str, business_timezone,
                total_duration,
            ),
        )
        time_row = cur.fetchone()
        start_time_db = time_row["start_time_db"]
        end_time_db = time_row["end_time_db"]

        overlap = find_overlapping_appointment(
            cur,
            business["id"],
            start_time_db,
            end_time_db,
            business_timezone,
            staff_id=staff_id,
            exclude_appointment_id=appointment_id,
        )
        if overlap:
            conn.rollback()
            flash(build_overlap_message(overlap, staff_id=staff_id), "warning")
            return redirect(url_for("nails.agenda"))

        cur.execute(
            """
            UPDATE nails_appointments
            SET client_id = %s,
                staff_id = %s,
                service_id = %s,
                title = %s,
                start_time = %s,
                end_time = %s,
                status = %s,
                estimated_total = %s,
                deposit_amount = %s,
                notes = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND business_id = %s
            """,
            (
                client_id if client_id else None,
                staff_id if staff_id else None,
                service_id,
                " + ".join(item["name"] for item in services_to_insert[:3]),
                start_time_db,
                end_time_db,
                status,
                estimated_total,
                parse_positive_float(deposit_amount_raw),
                notes,
                appointment_id,
                business["id"],
            ),
        )

        cur.execute("DELETE FROM nails_appointment_services WHERE appointment_id = %s", (appointment_id,))
        cur.execute("DELETE FROM nails_appointment_extras WHERE appointment_id = %s", (appointment_id,))

        for service_item in services_to_insert:
            cur.execute(
                """
                INSERT INTO nails_appointment_services (
                    appointment_id, service_id, name, price, duration_minutes, sort_order
                )
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (
                    appointment_id,
                    service_item["id"],
                    service_item["name"],
                    service_item["price"],
                    service_item["duration_minutes"],
                    service_item["sort_order"],
                ),
            )

        for extra in extras_to_insert:
            cur.execute(
                """
                INSERT INTO nails_appointment_extras (
                    appointment_id, extra_id, name, price, duration_minutes
                )
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    appointment_id,
                    extra["id"],
                    extra["name"],
                    extra["price"],
                    extra["duration_minutes"],
                ),
            )

        sale_detail_items = [
            {
                "item_type": "service",
                "item_id": item["id"],
                "name": item["name"],
                "description": None,
                "quantity": 1,
                "unit_price": item["price"],
                "total": item["price"],
            }
            for item in services_to_insert
        ]
        sale_detail_items.extend(
            {
                "item_type": "extra",
                "item_id": extra["id"],
                "name": extra["name"],
                "description": None,
                "quantity": 1,
                "unit_price": extra["price"],
                "total": extra["price"],
            }
            for extra in extras_to_insert
        )
        sync_sale_for_appointment(
            cur,
            business["id"],
            appointment_id,
            client_id,
            staff_id,
            sale_detail_items,
            estimated_total,
            initial_paid_amount=parse_positive_float(deposit_amount_raw),
            notes=notes,
        )

        if status == "cancelada":
            cur.execute(
                """
                UPDATE nails_sales
                SET status = 'cancelada',
                    balance_due = 0,
                    updated_at = CURRENT_TIMESTAMP
                WHERE appointment_id = %s
                  AND business_id = %s
                  AND status != 'cancelada'
                """,
                (appointment_id, business["id"]),
            )

        cur.execute(
            """
            INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (
                business["id"], user_id,
                "appointment_update", "Agenda",
                f"Editó la cita #{appointment_id}",
            ),
        )

        conn.commit()
        flash("Cita actualizada correctamente.", "success")
        return redirect(url_for("nails.agenda"))

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo editar la cita: {e}", "danger")
        return redirect(url_for("nails.agenda"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# ELIMINAR CITA
# POST: Acción destructiva reservada a la dueña/owner del salón.
# =========================================================

@nails_bp.route("/agenda/<int:appointment_id>/eliminar", methods=["POST"])
def eliminar_cita(appointment_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        if not user_is_nails_owner(cur, user_id, business):
            conn.rollback()
            flash("Solo la dueña del salón puede eliminar citas.", "warning")
            return redirect(url_for("nails.agenda"))

        cur.execute(
            """
            SELECT
                a.id,
                a.title,
                c.name AS client_name,
                TO_CHAR(a.start_time AT TIME ZONE %s, 'DD/MM/YYYY HH24:MI') AS start_time_label
            FROM nails_appointments a
            LEFT JOIN nails_clients c ON c.id = a.client_id
            WHERE a.id = %s AND a.business_id = %s
            LIMIT 1
            """,
            (business["timezone"] or "America/Monterrey", appointment_id, business["id"]),
        )
        appointment = cur.fetchone()

        if not appointment:
            conn.rollback()
            flash("No se encontró la cita.", "warning")
            return redirect(url_for("nails.agenda"))

        cur.execute(
            """
            UPDATE nails_sales
            SET status = 'cancelada',
                balance_due = 0,
                updated_at = CURRENT_TIMESTAMP
            WHERE appointment_id = %s
              AND business_id = %s
              AND status != 'cancelada'
            """,
            (appointment_id, business["id"]),
        )

        cur.execute("DELETE FROM nails_appointment_services WHERE appointment_id = %s", (appointment_id,))
        cur.execute("DELETE FROM nails_appointment_extras WHERE appointment_id = %s", (appointment_id,))
        cur.execute(
            """
            DELETE FROM nails_appointments
            WHERE id = %s AND business_id = %s
            """,
            (appointment_id, business["id"]),
        )

        cur.execute(
            """
            INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (
                business["id"], user_id,
                "appointment_delete", "Agenda",
                (
                    f"Eliminó la cita #{appointment_id}: "
                    f"{appointment['title'] or 'Sin título'} - "
                    f"{appointment['client_name'] or 'Cliente General'} - "
                    f"{appointment['start_time_label']}"
                ),
            ),
        )
        conn.commit()
        flash("Cita eliminada correctamente.", "success")
        return redirect(url_for("nails.agenda"))

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo eliminar la cita: {e}", "danger")
        return redirect(url_for("nails.agenda"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# CATÁLOGO PÚBLICO
# Vista pública accesible sin login mediante el slug del salón.
# Muestra servicios, extras, galería y horarios del negocio.
# =========================================================

@nails_bp.route("/catalogo/<slug>", methods=["GET", "POST"])
def catalogo_publico(slug):
    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        ensure_nails_service_icon_column(cur)
        conn.commit()

        cur.execute(
            """
            SELECT * FROM nails_businesses
            WHERE slug = %s AND is_active = TRUE
            LIMIT 1
            """,
            (slug,),
        )
        public_business = cur.fetchone()

        if not public_business:
            return "Catálogo no encontrado", 404

        if request.method == "POST":
            rate_key = (
                request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip(),
                public_business["id"],
            )
            if public_booking_rate_limited(rate_key):
                flash("Recibimos muchas solicitudes desde este dispositivo. Intenta de nuevo en unos minutos.", "warning")
                return redirect(url_for("nails.catalogo_publico", slug=slug, agendar="1"))

            ensure_nails_appointment_services_table(cur)

            client_name = clean_text(request.form.get("client_name"), 120)
            client_phone = clean_text(request.form.get("client_phone"), 40)
            selected_service_ids = [
                clean_optional_id(value)
                for value in request.form.getlist("service_ids")
            ]
            selected_service_ids = [value for value in selected_service_ids if value]
            service_id_from_picker = clean_optional_id(request.form.get("service_id"))
            if not selected_service_ids and service_id_from_picker:
                selected_service_ids = [service_id_from_picker]
            service_id = selected_service_ids[0] if selected_service_ids else None
            selected_extra_ids = [
                clean_optional_id(value)
                for value in request.form.getlist("extras")
            ]
            selected_extra_ids = [value for value in selected_extra_ids if value]
            appointment_date = clean_text(request.form.get("appointment_date"))
            start_time = clean_text(request.form.get("start_time"))
            notes = clean_text(request.form.get("notes"), 1000)

            if not client_name or not client_phone or not service_id or not appointment_date or not start_time:
                flash("Completa nombre, WhatsApp, servicio, fecha y hora para agendar.", "warning")
                return redirect(url_for("nails.catalogo_publico", slug=slug, agendar="1"))

            if not parse_date_value(appointment_date) or not re.fullmatch(r"\d{2}:\d{2}", start_time):
                flash("Fecha u hora inválida.", "warning")
                return redirect(url_for("nails.catalogo_publico", slug=slug, agendar="1"))

            available_slots, services_selected, extras_selected, slots_message = get_public_available_slots(
                cur,
                public_business,
                selected_service_ids,
                appointment_date,
                selected_extra_ids,
            )

            if not services_selected:
                flash("Uno o más servicios seleccionados ya no están disponibles.", "warning")
                return redirect(url_for("nails.catalogo_publico", slug=slug, agendar="1"))

            if start_time not in {slot["value"] for slot in available_slots}:
                flash(slots_message or "Ese horario ya no está disponible. Elige otra hora para agendar tu cita.", "warning")
                return redirect(url_for("nails.catalogo_publico", slug=slug, agendar="1"))

            appointment_date_value = parse_date_value(appointment_date)
            business_timezone = public_business["timezone"] or "America/Monterrey"
            start_datetime_str = f"{appointment_date} {start_time}:00"
            service_duration = sum(max(15, int(service["duration_minutes"] or 60)) for service in services_selected)
            service_duration += sum(max(0, int(extra["duration_minutes"] or 0)) for extra in extras_selected)
            estimated_total = sum(float(service["base_price"] or 0) for service in services_selected)
            estimated_total += sum(float(extra["price"] or 0) for extra in extras_selected)
            appointment_title = " + ".join(service["name"] for service in services_selected[:3])

            cur.execute(
                """
                SELECT
                    (%s::timestamp AT TIME ZONE %s) AS start_time_db,
                    ((%s::timestamp AT TIME ZONE %s)
                     + (%s || ' minutes')::interval) AS end_time_db,
                    ((%s::timestamp AT TIME ZONE %s) < NOW()) AS is_past
                """,
                (
                    start_datetime_str,
                    business_timezone,
                    start_datetime_str,
                    business_timezone,
                    service_duration,
                    start_datetime_str,
                    business_timezone,
                ),
            )
            time_row = cur.fetchone()

            if time_row["is_past"]:
                flash("Elige una fecha y hora futura para tu cita.", "warning")
                return redirect(url_for("nails.catalogo_publico", slug=slug, agendar="1"))

            overlap = find_overlapping_appointment(
                cur,
                public_business["id"],
                time_row["start_time_db"],
                time_row["end_time_db"],
                business_timezone,
                staff_id=None,
            )
            if overlap:
                flash("Ese horario ya no está disponible. Elige otra hora para agendar tu cita.", "warning")
                return redirect(url_for("nails.catalogo_publico", slug=slug, agendar="1"))

            client_id = None
            if not is_general_customer_name(client_name):
                cur.execute(
                    """
                    INSERT INTO nails_clients (business_id, name, phone, notes)
                    VALUES (%s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        public_business["id"],
                        client_name,
                        client_phone,
                        "Clienta creada desde catálogo público",
                    ),
                )
                client_id = cur.fetchone()["id"]

            cur.execute(
                """
                INSERT INTO nails_appointments (
                    business_id, client_id, staff_id, service_id,
                    title, start_time, end_time, status,
                    estimated_total, deposit_amount, notes
                )
                VALUES (%s, %s, NULL, %s, %s, %s, %s, 'pendiente', %s, 0, %s)
                RETURNING id
                """,
                (
                    public_business["id"],
                    client_id,
                    service_id,
                    appointment_title,
                    time_row["start_time_db"],
                    time_row["end_time_db"],
                    estimated_total,
                    notes,
                ),
            )
            appointment_id = cur.fetchone()["id"]

            for index, service in enumerate(services_selected):
                cur.execute(
                    """
                    INSERT INTO nails_appointment_services (
                        appointment_id, service_id, name, price, duration_minutes, sort_order
                    )
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (
                        appointment_id,
                        service["id"],
                        service["name"],
                        float(service["base_price"] or 0),
                        int(service["duration_minutes"] or 60),
                        index,
                    ),
                )

            for extra in extras_selected:
                cur.execute(
                    """
                    INSERT INTO nails_appointment_extras (
                        appointment_id, extra_id, name, price, duration_minutes
                    )
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        appointment_id,
                        extra["id"],
                        extra["name"],
                        float(extra["price"] or 0),
                        int(extra["duration_minutes"] or 0),
                    ),
                )

            sale_detail_items = [
                {
                    "item_type": "service",
                    "item_id": service["id"],
                    "name": service["name"],
                    "description": None,
                    "quantity": 1,
                    "unit_price": float(service["base_price"] or 0),
                    "total": float(service["base_price"] or 0),
                }
                for service in services_selected
            ]
            sale_detail_items.extend(
                {
                    "item_type": "extra",
                    "item_id": extra["id"],
                    "name": extra["name"],
                    "description": None,
                    "quantity": 1,
                    "unit_price": float(extra["price"] or 0),
                    "total": float(extra["price"] or 0),
                }
                for extra in extras_selected
            )

            sync_sale_for_appointment(
                cur,
                public_business["id"],
                appointment_id,
                client_id,
                None,
                sale_detail_items,
                estimated_total,
                initial_paid_amount=0,
                notes=notes,
            )

            cur.execute(
                """
                INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
                VALUES (%s, NULL, %s, %s, %s)
                """,
                (
                    public_business["id"],
                    "public_appointment_create",
                    "Catálogo público",
                    f"{client_name} agendó {appointment_title} para {appointment_date_value} {start_time}",
                ),
            )

            conn.commit()
            flash("Tu cita fue solicitada correctamente. El salón podrá confirmarla pronto.", "success")
            return redirect(url_for("nails.catalogo_publico", slug=slug, agendada="1"))

        # Solo servicios marcados como públicos
        cur.execute(
            """
            SELECT s.*, c.name AS category_name
            FROM nails_services s
            LEFT JOIN nails_service_categories c ON c.id = s.category_id
            WHERE s.business_id = %s
              AND s.is_active   = TRUE
              AND s.is_public   = TRUE
            ORDER BY c.sort_order ASC, s.sort_order ASC, s.name ASC
            """,
            (public_business["id"],),
        )
        services = cur.fetchall()

        cur.execute(
            """
            SELECT * FROM nails_extras
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY sort_order ASC, name ASC
            """,
            (public_business["id"],),
        )
        extras = cur.fetchall()

        # Parsear horarios de apertura desde JSON
        business_hours = {}
        if public_business["business_hours_json"]:
            try:
                business_hours = json.loads(public_business["business_hours_json"])
            except Exception:
                business_hours = {}

        # Últimas 12 imágenes de galería públicas
        cur.execute(
            """
            SELECT g.*, s.name AS service_name
            FROM nails_gallery g
            LEFT JOIN nails_services s ON s.id = g.service_id
            WHERE g.business_id = %s
              AND g.is_active   = TRUE
              AND g.is_public   = TRUE
            ORDER BY g.sort_order ASC, g.created_at DESC
            LIMIT 12
            """,
            (public_business["id"],),
        )
        gallery_items = cur.fetchall()

        return render_template(
            "nails/catalogo_publico.html",
            business=public_business,
            services=services,
            extras=extras,
            gallery_items=gallery_items,
            business_hours=business_hours,
        )

    finally:
        cur.close()
        conn.close()


@nails_bp.route("/catalogo/<slug>/horarios")
def catalogo_horarios_disponibles(slug):
    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            SELECT * FROM nails_businesses
            WHERE slug = %s AND is_active = TRUE
            LIMIT 1
            """,
            (slug,),
        )
        public_business = cur.fetchone()

        if not public_business:
            return jsonify({
                "success": False,
                "slots": [],
                "message": "Catálogo no encontrado.",
            }), 404

        service_ids = request.args.getlist("service_ids")
        if not service_ids:
            service_id = request.args.get("service_id")
            service_ids = [service_id] if service_id else []
        extra_ids = request.args.getlist("extras")
        appointment_date = clean_text(request.args.get("date"))
        slots, services_selected, extras_selected, message = get_public_available_slots(
            cur,
            public_business,
            service_ids,
            appointment_date,
            extra_ids,
        )
        total_duration = sum(int(service["duration_minutes"] or 60) for service in services_selected)
        total_duration += sum(int(extra["duration_minutes"] or 0) for extra in extras_selected)
        total_price = sum(float(service["base_price"] or 0) for service in services_selected)
        total_price += sum(float(extra["price"] or 0) for extra in extras_selected)

        return jsonify({
            "success": True,
            "slots": slots,
            "message": message,
            "services": [
                {
                    "id": service["id"],
                    "name": service["name"],
                    "duration_minutes": service["duration_minutes"],
                    "base_price": float(service["base_price"] or 0),
                }
                for service in services_selected
            ],
            "extras": [
                {
                    "id": extra["id"],
                    "name": extra["name"],
                    "duration_minutes": extra["duration_minutes"],
                    "price": float(extra["price"] or 0),
                }
                for extra in extras_selected
            ],
            "summary": {
                "duration_minutes": total_duration,
                "total": total_price,
            },
        })

    finally:
        cur.close()
        conn.close()


# =========================================================
# UPLOAD A R2 (ENDPOINT AJAX)
# Recibe un archivo de imagen vía multipart/form-data,
# lo sube a Cloudflare R2 y devuelve la URL pública.
# Si el folder_type es 'logo', actualiza también logo_url
# en la BD y registra el evento en el log de actividad.
# =========================================================

@nails_bp.route("/upload-r2", methods=["POST"])
def upload_r2_nails():
    if not require_login():
        return jsonify({"success": False, "error": "No autorizado"}), 401

    user_id  = session.get("user_id")
    username = session.get("username", "Anonimo")
    business = get_user_nails_business(user_id)

    if not business:
        return jsonify({"success": False, "error": "No se encontró negocio Nails"}), 404

    file = request.files.get("file")

    if not file:
        return jsonify({"success": False, "error": "No se recibió ningún archivo"}), 400

    if not allowed_image_file(file.filename or ""):
        return jsonify({"success": False, "error": "Formato de imagen no permitido"}), 400

    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)

    if file_size > 8 * 1024 * 1024:
        return jsonify({"success": False, "error": "La imagen no puede pesar más de 8 MB"}), 400

    # Validar tipo de carpeta destino contra whitelist
    folder_type = request.form.get("folder_type", "general").strip()
    allowed_folder_types = {
        "logo":    "logo",
        "gallery": "gallery",
        "service": "services",
        "general": "general",
    }
    folder_name = allowed_folder_types.get(folder_type, "general")

    mime_type = file.content_type or ""
    if not mime_type.startswith("image/"):
        return jsonify({"success": False, "error": "Solo se permiten imágenes"}), 400

    base_filename   = secure_filename(file.filename or "imagen")
    unique_filename = f"{uuid.uuid4().hex}_{base_filename}"
    ruta_r2         = f"nails/business_{business['id']}/{folder_name}/{unique_filename}"

    try:
        get_s3_client().upload_fileobj(
            file,
            BUCKET_NAME,
            ruta_r2,
            ExtraArgs={"ContentType": mime_type},
        )
        url_final = f"{PUBLIC_URL}/{ruta_r2}"

        # Si es logo, actualizar la BD además de subir el archivo
        if folder_type == "logo":
            conn = get_db_connection()
            cur  = conn.cursor()
            try:
                cur.execute(
                    """
                    UPDATE nails_businesses
                    SET logo_url   = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id      = %s
                      AND user_id = %s
                    """,
                    (url_final, business["id"], user_id),
                )
                cur.execute(
                    """
                    INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        business["id"], user_id,
                        "business_logo_upload", "Configuración",
                        "Actualizó el logo del salón",
                    ),
                )
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            finally:
                cur.close()
                conn.close()

        current_app.logger.info(
            f"NAILS_R2_UPLOAD_SUCCESS: Usuario '{username}' ID {user_id} subió '{ruta_r2}'"
        )
        return jsonify({"success": True, "url": url_final})

    except Exception as e:
        current_app.logger.error(
            f"NAILS_R2_UPLOAD_ERROR: Usuario '{username}' ID {user_id} falló al subir archivo - {e}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


# =========================================================
# GALERÍA (ADMIN)
# GET:  Lista todas las imágenes de galería del salón (activas).
# POST: Agrega una nueva imagen a la galería usando la URL
#       devuelta por /upload-r2 (flujo en dos pasos).
# =========================================================

@nails_bp.route("/galeria", methods=["GET", "POST"])
def galeria():
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        if request.method == "POST":
            title      = clean_text(request.form.get("title"), 160)
            description = clean_text(request.form.get("description"), 1200)
            image_url  = clean_text(request.form.get("image_url"), 1000)
            service_id = clean_optional_id(request.form.get("service_id"))
            sort_order = parse_positive_int(request.form.get("sort_order"), default=0, max_value=100000)
            is_public  = request.form.get("is_public") == "on"

            if not image_url:
                flash("Primero sube una imagen.", "warning")
                return redirect(url_for("nails.galeria"))

            if service_id and not row_belongs_to_business(cur, "nails_services", service_id, business["id"]):
                flash("El servicio relacionado no pertenece a este salón.", "warning")
                return redirect(url_for("nails.galeria"))

            if not title:
                title = "Trabajo realizado"

            cur.execute(
                """
                INSERT INTO nails_gallery (
                    business_id, service_id, title, description,
                    image_url, is_public, sort_order
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    business["id"], service_id, title, description,
                    image_url, is_public, sort_order,
                ),
            )
            cur.execute(
                """
                INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    business["id"], user_id,
                    "gallery_image_create", "Galería",
                    f"Agregó imagen a galería: {title}",
                ),
            )
            conn.commit()
            flash("Imagen agregada a la galería.", "success")
            return redirect(url_for("nails.galeria"))

        # ── GET: cargar servicios e imágenes ───────────────
        cur.execute(
            """
            SELECT id, name FROM nails_services
            WHERE business_id = %s AND is_active = TRUE
            ORDER BY name ASC
            """,
            (business["id"],),
        )
        services = cur.fetchall()

        cur.execute(
            """
            SELECT g.*, s.name AS service_name
            FROM nails_gallery g
            LEFT JOIN nails_services s ON s.id = g.service_id
            WHERE g.business_id = %s
            ORDER BY g.sort_order ASC, g.created_at DESC
            """,
            (business["id"],),
        )
        gallery_items = cur.fetchall()

        gallery_active_items = [item for item in gallery_items if item["is_active"]]
        gallery_public_count = sum(1 for item in gallery_active_items if item["is_public"])
        gallery_hidden_count = len(gallery_active_items) - gallery_public_count
        services_with_photos = len({
            item["service_id"]
            for item in gallery_active_items
            if item["service_id"]
        })

        return render_template(
            "nails/galeria.html",
            business=business,
            services=services,
            gallery_items=gallery_items,
            gallery_stats={
                "total": len(gallery_active_items),
                "public": gallery_public_count,
                "hidden": gallery_hidden_count,
                "services_with_photos": services_with_photos,
            },
        )

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo cargar galería: {e}", "danger")
        return redirect(url_for("nails.dashboard"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# ELIMINAR IMAGEN DE GALERÍA (soft delete)
# Marca la imagen como is_active=FALSE en lugar de borrarla.
# =========================================================

@nails_bp.route("/galeria/<int:image_id>/eliminar", methods=["POST"])
def eliminar_imagen_galeria(image_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            UPDATE nails_gallery
            SET is_active  = FALSE,
                updated_at = CURRENT_TIMESTAMP
            WHERE id          = %s
              AND business_id = %s
            RETURNING title
            """,
            (image_id, business["id"]),
        )
        deleted = cur.fetchone()

        if not deleted:
            conn.rollback()
            flash("No se encontró la imagen.", "warning")
            return redirect(url_for("nails.galeria"))

        cur.execute(
            """
            INSERT INTO nails_activity_logs (business_id, user_id, action, module, detail)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (
                business["id"], user_id,
                "gallery_image_delete", "Galería",
                f"Ocultó imagen de galería: {deleted['title']}",
            ),
        )
        conn.commit()
        flash("Imagen eliminada de la galería.", "success")
        return redirect(url_for("nails.galeria"))

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo eliminar la imagen: {e}", "danger")
        return redirect(url_for("nails.galeria"))

    finally:
        cur.close()
        conn.close()


# =========================================================
# TOGGLE VISIBILIDAD EN CATÁLOGO PÚBLICO
# Alterna el campo is_public de una imagen de galería.
# =========================================================

@nails_bp.route("/galeria/<int:image_id>/toggle-public", methods=["POST"])
def toggle_public_galeria(image_id):
    if not require_login():
        return redirect(url_for("auth.login"))

    user_id  = session.get("user_id")
    business = get_user_nails_business(user_id)

    if not business:
        return redirect(url_for("nails.onboarding"))

    conn = get_db_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            UPDATE nails_gallery
            SET is_public  = NOT is_public,
                updated_at = CURRENT_TIMESTAMP
            WHERE id          = %s
              AND business_id = %s
              AND is_active   = TRUE
            RETURNING is_public
            """,
            (image_id, business["id"]),
        )
        updated = cur.fetchone()

        if not updated:
            conn.rollback()
            flash("No se encontró la imagen.", "warning")
            return redirect(url_for("nails.galeria"))

        conn.commit()
        flash(
            "Imagen visible en catálogo." if updated["is_public"] else "Imagen oculta del catálogo.",
            "success" if updated["is_public"] else "info",
        )
        return redirect(url_for("nails.galeria"))

    except Exception as e:
        conn.rollback()
        flash(f"No se pudo actualizar la imagen: {e}", "danger")
        return redirect(url_for("nails.galeria"))

    finally:
        cur.close()
        conn.close()
